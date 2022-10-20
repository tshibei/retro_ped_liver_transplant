from openpyxl import load_workbook
from scipy.optimize import curve_fit
import math
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.pylab as pylab
import matplotlib.patches as patches
import numpy as np
from scipy import stats
import seaborn as sns
from statistics import mean
import sys
pd.options.mode.chained_assignment = None  # default='warn'

# CURATE
def execute_CURATE(five_fold_cross_val_results_summary='', pop_tau_string='', dose='total'):
    """ 
    Execute CURATE.

    Output: 
    'CURATE_results.xlsx' if dose is 'total', and 'CURATE_results_evening_dose.xlsx' if dose is 'evening'
            Excel sheet with cleaned patient dataframe, 
            dataframe for calibration and efficacy-driven dosing, 
            result of all methods.
    """
    # Generate profiles and join dataframes
    patients_to_exclude_linear, patients_to_exclude_quad, list_of_patient_df, list_of_cal_pred_df, list_of_result_df = generate_profiles(five_fold_cross_val_results_summary, dose)
    df, cal_pred, result_df = join_dataframes(list_of_patient_df, list_of_cal_pred_df, list_of_result_df)

    # Print patients to exclude anad ouput dataframes to excel as individual sheets
    print_patients_to_exclude(patients_to_exclude_linear, patients_to_exclude_quad)
    output_df_to_excel(df, cal_pred, result_df, pop_tau_string, dose)

# Pop tau
def find_pop_tau_with_CV():
    """
    Calculate pop tau with five fold cross validation.

    Details:
    For each experiment, find the half life with lowest absolute deviation in the training set, and use that half life to
    find the absolute deviation of both training and test sets.

    Output: 
    - five_fold_cross_val_results: dataframe of results of cross validation per experiment
    - five_fold_cross_val_results_summary: dataframe of results of cross validation per method
    """
    dat = pd.read_excel('output (no pop tau).xlsx', sheet_name='result')

    # Filter for tau methods only
    dat = dat[dat.method.str.contains('tau')]

    # Create patient and method list
    linear_patient_list = dat[dat.method.str.contains('L_')].patient.unique().tolist()
    quad_patient_list = dat[dat.method.str.contains('Q_')].patient.unique().tolist()
    method_list = dat.method.unique().tolist()

    # Calculate mean abs deviation by grouping by method, patient, half-life 
    dat = dat.groupby(['method', 'patient', 'half_life'])['abs_deviation'].mean()
    dat = dat.to_frame(name='abs_deviation').reset_index()

    # Run normality check (result p=0.0, reject normality)
    # print(f'KS test for abs deviation: {scipy.stats.kstest(dat.abs_deviation, \'norm')}')

    # Define each fold
    linear_fold_1 = linear_patient_list[0:2]
    linear_fold_2 = linear_patient_list[2:4]
    linear_fold_3 = linear_patient_list[4:7]
    linear_fold_4 = linear_patient_list[7:10]
    linear_fold_5 = linear_patient_list[10:]

    quad_fold_1 = quad_patient_list[0:2]
    quad_fold_2 = quad_patient_list[2:4]
    quad_fold_3 = quad_patient_list[4:6]
    quad_fold_4 = quad_patient_list[6:8]
    quad_fold_5 = quad_patient_list[8:]

    # Define each experiment
    linear_exp_1_train = linear_fold_1 + linear_fold_2 + linear_fold_3 + linear_fold_4
    linear_exp_2_train = linear_fold_1 + linear_fold_2 + linear_fold_3 + linear_fold_5
    linear_exp_3_train = linear_fold_1 + linear_fold_2 + linear_fold_4 + linear_fold_5
    linear_exp_4_train = linear_fold_1 + linear_fold_3 + linear_fold_4 + linear_fold_5
    linear_exp_5_train = linear_fold_2 + linear_fold_3 + linear_fold_4 + linear_fold_5

    quad_exp_1_train = quad_fold_1 + quad_fold_2 + quad_fold_3 + quad_fold_4
    quad_exp_2_train = quad_fold_1 + quad_fold_2 + quad_fold_3 + quad_fold_5
    quad_exp_3_train = quad_fold_1 + quad_fold_2 + quad_fold_4 + quad_fold_5
    quad_exp_4_train = quad_fold_1 + quad_fold_3 + quad_fold_4 + quad_fold_5
    quad_exp_5_train = quad_fold_2 + quad_fold_3 + quad_fold_4 + quad_fold_5

    linear_exp_1_test = linear_fold_5
    linear_exp_2_test = linear_fold_4
    linear_exp_3_test = linear_fold_3
    linear_exp_4_test = linear_fold_2
    linear_exp_5_test = linear_fold_1

    quad_exp_1_test = quad_fold_5
    quad_exp_2_test = quad_fold_4
    quad_exp_3_test = quad_fold_3
    quad_exp_4_test = quad_fold_2
    quad_exp_5_test = quad_fold_1

    list_of_linear_train = [linear_exp_1_train, linear_exp_2_train, linear_exp_3_train, linear_exp_4_train, linear_exp_5_train]
    list_of_quad_train = [quad_exp_1_train, quad_exp_2_train, quad_exp_3_train, quad_exp_4_train, quad_exp_5_train]
    list_of_linear_test = [linear_exp_1_test, linear_exp_2_test, linear_exp_3_test, linear_exp_4_test, linear_exp_5_test]
    list_of_quad_test = [quad_exp_1_test, quad_exp_2_test, quad_exp_3_test, quad_exp_4_test, quad_exp_5_test]

    # Define dataframe to store results
    five_fold_cross_val_results = pd.DataFrame(columns=['method', 'experiment', 'train_median', 'test_median', 'pop_half_life_fold', 'indiv_pop_half_life_fold'])
    five_fold_cross_val_results_summary = pd.DataFrame(columns=['method', 'train_median_mean', 'train_median_SEM', \
                                                                'test_median_mean', 'test_median_SEM', \
                                                                'pop_half_life'])
    fold_counter = 1
    method_counter = 1

    for method in method_list: # loop through methods
        list_of_pop_half_life_fold = []
        list_of_median_abs_dev_train = []
        list_of_median_abs_dev_test = []
        method_df = dat[dat.method == method]

        for i in range(5): # loop through experiments

            # Define train_df
            if 'L_' in method:
                train_df = method_df[method_df.patient.isin(list_of_linear_train[i])]
            else:
                train_df = method_df[method_df.patient.isin(list_of_quad_train[i])]

            # Find half_life at the index where abs_deviation is the lowest
            train_df.reset_index()
            pop_half_life_fold_index = train_df.index[train_df.abs_deviation == train_df.abs_deviation.min()].tolist()
            pop_half_life_fold = train_df.loc[pop_half_life_fold_index, 'half_life'].tolist()

            # Find median of abs_deviation among train_df with pop_half_life_fold
            median_abs_dev_train = round(train_df[train_df.half_life.isin(pop_half_life_fold)].abs_deviation.median(), 2)

            # Define test_df
            if 'L_' in method:
                test_df = method_df[method_df.patient.isin(list_of_linear_test[i])]
            else:
                test_df = method_df[method_df.patient.isin(list_of_quad_test[i])]

            # Find median of abs_deviation among test_df with pop_half_life_fold
            median_abs_dev_test = round(test_df[test_df.half_life.isin(pop_half_life_fold)].abs_deviation.median(), 2)

            # If there are multiple half-lives with the lowest abs deviation, find average of the half lives to store as pop_half_life_fold
            if len(pop_half_life_fold) > 1:
                indiv_pop_half_life_fold = pop_half_life_fold
                pop_half_life_fold = mean(pop_half_life_fold)
                pop_half_life_fold = [pop_half_life_fold]
            else: 
                pop_half_life_fold = pop_half_life_fold
                indiv_pop_half_life_fold = ""

            pop_half_life_fold = pop_half_life_fold[0]

            list_of_pop_half_life_fold.append(pop_half_life_fold)
            list_of_median_abs_dev_train.append(median_abs_dev_train)
            list_of_median_abs_dev_test.append(median_abs_dev_test)

            # Fill in five_fold_cross_val_results for results per fold
            five_fold_cross_val_results.loc[fold_counter, 'method'] = method
            five_fold_cross_val_results.loc[fold_counter, 'experiment'] = i + 1
            five_fold_cross_val_results.loc[fold_counter, 'train_median'] = median_abs_dev_train
            five_fold_cross_val_results.loc[fold_counter, 'test_median'] = median_abs_dev_test
            five_fold_cross_val_results.loc[fold_counter, 'pop_half_life_fold'] = pop_half_life_fold
            five_fold_cross_val_results.loc[fold_counter, 'indiv_pop_half_life_fold'] = indiv_pop_half_life_fold

            fold_counter = fold_counter + 1

        # Fill in five_fold_cross_val_results_summary for results per method
        five_fold_cross_val_results_summary.loc[fold_counter, 'method'] = method
        five_fold_cross_val_results_summary.loc[fold_counter, 'train_median_mean'] = round(mean(list_of_median_abs_dev_train), 2)
        five_fold_cross_val_results_summary.loc[fold_counter, 'train_median_SEM'] = round(stats.sem(list_of_median_abs_dev_train), 2)
        five_fold_cross_val_results_summary.loc[fold_counter, 'test_median_mean'] = round(mean(list_of_median_abs_dev_test), 2)
        five_fold_cross_val_results_summary.loc[fold_counter, 'test_median_SEM'] = round(stats.sem(list_of_median_abs_dev_test), 2)
        five_fold_cross_val_results_summary.loc[fold_counter, 'pop_half_life'] = sum(list_of_pop_half_life_fold) / len(list_of_pop_half_life_fold)

        method_counter = method_counter + 1

    five_fold_cross_val_results = five_fold_cross_val_results.reset_index(drop=True)
    five_fold_cross_val_results_summary = five_fold_cross_val_results_summary.reset_index(drop=True)
    
    return five_fold_cross_val_results, five_fold_cross_val_results_summary

def find_pop_tau_with_LOOCV():
    """
    Calculate pop tau with leave one out cross validation (LOOCV).

    Details:
    For each experiment, find the half life with lowest absolute deviation in the training set, and use that half life to
    find the absolute deviation of both training and test sets.

    Output: 
    - five_fold_cross_val_results: dataframe of results of LOOCV per experiment
    - five_fold_cross_val_results_summary: dataframe of results of LOOCV per method
    """

    dat = pd.read_excel('output (no pop tau).xlsx', sheet_name='result')

    # Filter for tau methods only
    dat = dat[dat.method.str.contains('tau')]

    # Create patient and method list
    linear_patient_list = dat[dat.method.str.contains('L_')].patient.unique().tolist()
    quad_patient_list = dat[dat.method.str.contains('Q_')].patient.unique().tolist()
    method_list = dat.method.unique().tolist()

    # Calculate mean abs deviation by grouping by method, patient, half-life 
    dat = dat.groupby(['method', 'patient', 'half_life'])['abs_deviation'].mean()
    dat = dat.to_frame(name='abs_deviation').reset_index()

    # # Run normality check (result p=0.0, reject normality)
    # scipy.stats.kstest(dat.abs_deviation, 'norm')

    # Define dataframe to store results
    five_fold_cross_val_results = pd.DataFrame(columns=['method', 'experiment', 'train_median', 'test_median', 'pop_half_life_fold', 'indiv_pop_half_life_fold'])
    five_fold_cross_val_results_summary = pd.DataFrame(columns=['method', 'train_median_mean', 'train_median_SEM', \
                                                                'test_median_mean', 'test_median_SEM', \
                                                                'pop_half_life'])

    method_counter = 1
    fold_counter = 1

    for method in method_list: # loop through methods
        list_of_pop_half_life_fold = []
        list_of_median_abs_dev_train = []
        list_of_median_abs_dev_test = []
        method_df = dat[dat.method == method]

        # Check if linear/quad method

        # If linear, loop through 13 times
        if 'L_' in method:

            list_of_pop_half_life_fold, list_of_median_abs_dev_train, list_of_median_abs_dev_test, five_fold_cross_val_results, fold_counter = \
            LOOCV(linear_patient_list, method_df, list_of_pop_half_life_fold, list_of_median_abs_dev_train, \
                list_of_median_abs_dev_test, five_fold_cross_val_results, fold_counter, method)

        # If quad, loop through 11 times
        if 'Q_' in method:

            list_of_pop_half_life_fold, list_of_median_abs_dev_train, list_of_median_abs_dev_test, five_fold_cross_val_results, fold_counter = \
            LOOCV(quad_patient_list, method_df, list_of_pop_half_life_fold, list_of_median_abs_dev_train, \
                list_of_median_abs_dev_test, five_fold_cross_val_results, fold_counter, method)

        # Fill in five_fold_cross_val_results_summary for results per method
        five_fold_cross_val_results_summary.loc[fold_counter, 'method'] = method
        five_fold_cross_val_results_summary.loc[fold_counter, 'train_median_mean'] = round(mean(list_of_median_abs_dev_train), 2)
        five_fold_cross_val_results_summary.loc[fold_counter, 'train_median_SEM'] = round(stats.sem(list_of_median_abs_dev_train), 2)
        five_fold_cross_val_results_summary.loc[fold_counter, 'test_median_mean'] = round(mean(list_of_median_abs_dev_test), 2)
        five_fold_cross_val_results_summary.loc[fold_counter, 'test_median_SEM'] = round(stats.sem(list_of_median_abs_dev_test), 2)
        five_fold_cross_val_results_summary.loc[fold_counter, 'pop_half_life'] = sum(list_of_pop_half_life_fold) / len(list_of_pop_half_life_fold)

        method_counter = method_counter + 1

    five_fold_cross_val_results = five_fold_cross_val_results.reset_index(drop=True)
    five_fold_cross_val_results_summary = five_fold_cross_val_results_summary.reset_index(drop=True)

    return five_fold_cross_val_results, five_fold_cross_val_results_summary

def LOOCV(patient_list, method_df, list_of_pop_half_life_fold, list_of_median_abs_dev_train, list_of_median_abs_dev_test, \
    five_fold_cross_val_results, fold_counter, method):
    """
    LOOCV method for both linear and quadratic
    
    Output:
    - list_of_pop_half_life_fold: list of half-lives with the lowest absolute deviation in training set of that experiment
    - list_of_median_abs_dev_train: list of medians of absolute deviation of the training set
    - list_of_median_abs_dev_test: list of medians of absolute deviation of the test set
    - five_fold_cross_val_results: dataframe of results of cross validation per experiment
    - fold_counter: counter of number of folds thus far
    - method: name of method
    """
    for i in range(len(patient_list)):

        # Define test_df (i training dataset)
        test_df = method_df[method_df.patient == patient_list[i]]

        # Define train_df (pop i)
        train_list = patient_list.copy()
        removed_element = train_list.pop(i)
        train_df = method_df[method_df.patient.isin(train_list)]

        # Find half_life at the index where abs_deviation is the lowest
        train_df.reset_index()
        pop_half_life_fold_index = train_df.index[train_df.abs_deviation == train_df.abs_deviation.min()].tolist()
        pop_half_life_fold = train_df.loc[pop_half_life_fold_index, 'half_life'].tolist()

        # Find median of abs_deviation among train_df with pop_half_life_fold
        median_abs_dev_train = round(train_df[train_df.half_life.isin(pop_half_life_fold)].abs_deviation.median(), 2)

        # Find median of abs_deviation among test_df with pop_half_life_fold
        median_abs_dev_test = round(test_df[test_df.half_life.isin(pop_half_life_fold)].abs_deviation.median(), 2)

        # If there are multiple half-lives with the lowest abs deviation, find average of the half lives to store as pop_half_life_fold
        if len(pop_half_life_fold) > 1:
            indiv_pop_half_life_fold = pop_half_life_fold
            pop_half_life_fold = mean(pop_half_life_fold)
            pop_half_life_fold = [pop_half_life_fold]
        else: 
            pop_half_life_fold = pop_half_life_fold
            indiv_pop_half_life_fold = ""

        pop_half_life_fold = pop_half_life_fold[0]

        list_of_pop_half_life_fold.append(pop_half_life_fold)
        list_of_median_abs_dev_train.append(median_abs_dev_train)
        list_of_median_abs_dev_test.append(median_abs_dev_test)

        # Fill in five_fold_cross_val_results for results per fold
        five_fold_cross_val_results.loc[fold_counter, 'method'] = method
        five_fold_cross_val_results.loc[fold_counter, 'experiment'] = i + 1
        five_fold_cross_val_results.loc[fold_counter, 'train_median'] = median_abs_dev_train
        five_fold_cross_val_results.loc[fold_counter, 'test_median'] = median_abs_dev_test
        five_fold_cross_val_results.loc[fold_counter, 'pop_half_life_fold'] = pop_half_life_fold
        five_fold_cross_val_results.loc[fold_counter, 'indiv_pop_half_life_fold'] = indiv_pop_half_life_fold

        fold_counter = fold_counter + 1

    return list_of_pop_half_life_fold, list_of_median_abs_dev_train, list_of_median_abs_dev_test, five_fold_cross_val_results, fold_counter

def execute_CURATE_and_update_pop_tau_results(CV_string, five_fold_cross_val_results_summary, five_fold_cross_val_results):
    """
    Execute CURATE with pop tau and update pop tau results

    Output:
    Excel sheet containing dataframes for results of each experiment in 'Experiments' sheet and of overall results
    of each method in 'Overall' sheet.
    """
    five_fold_cross_val_results, five_fold_cross_val_results_summary = find_pop_tau(dose, cross_val_method)

    execute_CURATE(five_fold_cross_val_results_summary, pop_tau_string=' (of pop tau models only using ' + cross_val_method + ')', dose=dose)

    # Add pop_tau_method column
    five_fold_cross_val_results_summary['pop_tau_method'] = ""
    for i in range(len(five_fold_cross_val_results_summary)):
        five_fold_cross_val_results_summary.pop_tau_method[i] = five_fold_cross_val_results_summary.method[i][:-3] + 'pop_tau'

    # Import output with pop tau
    pop_tau_string=' (of pop tau models only using ' + cross_val_method + ')'
    if dose == 'evening':
        file_name = 'CURATE_results_evening_dose'+ pop_tau_string + '.xlsx'
    else:
        file_name = 'CURATE_results' + pop_tau_string + '.xlsx'

    pop_tau_df = pd.read_excel(file_name, sheet_name='result')

    # Filter pop tau methods
    pop_tau_df = pop_tau_df[pop_tau_df.method.str.contains('pop_tau')]

    # Calculate mean abs deviation by grouping by method, patient
    pop_tau_df = pop_tau_df.groupby(['method', 'patient'])['abs_deviation'].mean()
    pop_tau_df = pop_tau_df.to_frame(name='abs_deviation').reset_index()

    # Calculate median of 'mean of abs_deviation by patient' for each method
    pop_tau_df = pop_tau_df.groupby('method')['abs_deviation'].median()
    pop_tau_df = pop_tau_df.to_frame(name='abs_deviation').reset_index()

    # Rename pop_tau_df columns
    pop_tau_df.columns = ['pop_tau_method', 'all_patient_median']

    # Merge dataframes on 'pop_tau_method' column
    summary_df = five_fold_cross_val_results_summary.merge(pop_tau_df, how='left', on='pop_tau_method')

    # Output dataframes to excel as individual sheets
    with pd.ExcelWriter('experiments_to_find_pop_tau (by ' + cross_val_method + ').xlsx') as writer:
        five_fold_cross_val_results.to_excel(writer, sheet_name='Experiments', index=False)
        summary_df.to_excel(writer, sheet_name='Overall', index=False)

# Generate profiles
def generate_profiles(five_fold_cross_val_results_summary, dose):
    """
    Generate profiles for patients.
    
    Details:
    Define parameters and lists. Loop through patients to clean data, select data for calibration and efficacy-driven dosing, keep patients with
    sufficient dose-response pairs and predictions for each method, then apply method.
    
    Outputs:
    - patients_to_exclude_linear: list of patients to exclude for linear methods.
    - patients_to_exclude_quad: list of patients to exclude for quadratic methods.
    - list_of_patient_df: list of dataframes of cleaned patient data.
    - list_of_cal_pred_df: list of dataframes of data for calibration and efficacy-driven dosing, that has sufficient dose-response pairs and predictions
    - list_of_result_df: list of dataframes with results after applying methods.
    """

    # Profile Generation
    input_file = 'Retrospective Liver Transplant Data - edited.xlsx'
    rows_to_skip = 17

    # Get list of patients/sheet names
    list_of_patients = get_sheet_names(input_file)

    # Define lists
    list_of_patient_df = []
    list_of_cal_pred_df = []
    list_of_result_df = []
    patients_to_exclude_linear = []
    patients_to_exclude_quad = []
    list_of_body_weight = []

    # Create list of body_weight
    for i in range(len(list_of_patients)):    
        data = pd.read_excel(input_file, list_of_patients[i], index_col=None, usecols = "C", nrows=15)
        data = data.reset_index(drop=True)
        list_of_body_weight.append(data['Unnamed: 2'][13])
        
    list_of_body_weight = list_of_body_weight[:12]+[8.29]+list_of_body_weight[12+1:]

    number_of_patients = 0

    for patient in list_of_patients:

        # Create and clean patient dataframe        
        df = pd.read_excel(input_file, sheet_name=patient, skiprows=rows_to_skip)
        df = clean_data(df, dose)
        df = keep_ideal_data(df, patient, list_of_patient_df, dose)

        # Change to dose by body weight
        df['dose_BW'] = df['dose'] / list_of_body_weight[number_of_patients]
        
        # Counter for number of patients
        number_of_patients = number_of_patients + 1

        # Select data for calibration and efficacy-driven dosing
        cal_pred_linear, patients_to_exclude_linear = cal_pred_data(df, patient, patients_to_exclude_linear, 1)
        cal_pred_quad, patients_to_exclude_quad = cal_pred_data(df, patient, patients_to_exclude_quad, 2)

        # Keep patients with sufficient dose-response pairs and predictions for each method
        cal_pred, list_of_cal_pred_df = keep_target_patients(patient, patients_to_exclude_linear, patients_to_exclude_quad, 
                                                         cal_pred_linear, cal_pred_quad, list_of_cal_pred_df)   

        # Apply methods
        list_of_result_df = apply_methods(cal_pred, patient, patients_to_exclude_linear, patients_to_exclude_quad,
                  cal_pred_linear, cal_pred_quad, list_of_result_df, five_fold_cross_val_results_summary)

    return patients_to_exclude_linear, patients_to_exclude_quad, list_of_patient_df, list_of_cal_pred_df, list_of_result_df

def print_patients_to_exclude(patients_to_exclude_linear, patients_to_exclude_quad):
    """Print patients to exclude"""        
    patients_to_exclude_linear = sorted(set(patients_to_exclude_linear))
    patients_to_exclude_quad = sorted(set(patients_to_exclude_quad))
    print(f"Patients to exclude for linear methods: {patients_to_exclude_linear}")
    print(f"Patients to exclude for quad methods: {patients_to_exclude_quad}")

def join_dataframes(list_of_patient_df, list_of_cal_pred_df, list_of_result_df):
    """Join dataframes from individual patients"""
    df = pd.concat(list_of_patient_df)
    df.patient = df.patient.apply(int)
    df.reset_index(inplace=True, drop=True)
    cal_pred = pd.concat(list_of_cal_pred_df)
    cal_pred.patient = cal_pred.patient.apply(int)
    result_df = pd.concat(list_of_result_df)
    result_df = format_result_df(cal_pred, result_df)
    
    return df, cal_pred, result_df

def output_df_to_excel(df, cal_pred, result_df, pop_tau_string, dose):
    """Output dataframes to excel as individual sheets"""
    if dose == 'evening':
        file_name = 'CURATE_results_evening_dose'+ pop_tau_string + '.xlsx'
    else:
        file_name = 'CURATE_results' + pop_tau_string + '.xlsx'

    with pd.ExcelWriter(file_name) as writer:
        df.to_excel(writer, sheet_name='clean', index=False)
        cal_pred.to_excel(writer, sheet_name='calibration_and_efficacy_driven', index=False)
        result_df.to_excel(writer, sheet_name='result', index=False)

# Create patient dataframe        
def get_sheet_names(input_file):
    """ Get sheet names which are also patient names """
    wb = load_workbook(input_file, read_only=True)
    patient_list = wb.sheetnames
    wb.close()
    return patient_list

def clean_data(df, dose):
    """ 
    Keep target columns from excel, shift tac level one cell up, remove "mg"/"ng" 
    from dose, replace NaN with 0 in dose.
    
    Input:
    df - dataframe from each sheet
    patient - patient number from list of patients
    
    Output:
    df - cleaned dataframe        
    """
    if dose == 'total':
        dose_string = "Eff 24h Tac Dose"
    else:
        dose_string = "2nd Tac dose (pm)"

    # Keep target columns
    df = df[["Day #", "Tac level (prior to am dose)", dose_string]]

    # Shift tac level one cell up to match dose-response to one day
    df[dose_string] = df[dose_string].shift(1)

    # Remove "mg"/"ng" from dose
    df[dose_string] = df[dose_string].astype(str).str.replace('mgq', '')
    df[dose_string] = df[dose_string].astype(str).str.replace('mg', '')
    df[dose_string] = df[dose_string].astype(str).str.replace('ng', '')
    df[dose_string] = df[dose_string].astype(str).str.strip()
    df[dose_string] = df[dose_string].astype(float)

    first_day_of_dosing = df['Day #'].loc[~df[dose_string].isnull()].iloc[0]

    # Keep data from first day of dosing
    df = df[df['Day #'] >= first_day_of_dosing].reset_index(drop=True)

    # Set the first day of dosing as day 2 (because first dose received on day 1)
    for i in range(len(df)):
        df.loc[i, 'Day #'] = i + 2
    
    return df

def keep_ideal_data(df, patient, list_of_patient_df, dose):
    """
    Remove rows with non-ideal data, including missing tac level and/or tac dose, 
    <2 tac level, multiple blood draws. Then keep the longest consecutive chunk
    based on number of days. 

    Input: Dataframe of individual patient
    Output: Dataframe with the longest chunk of consecutive ideal data.
    """

    if dose == 'total':
        dose_string = "Eff 24h Tac Dose"
    else:
        dose_string = "2nd Tac dose (pm)"

    # Create boolean column of data to remove
    # Including NA, <2 tac level, multiple blood draws
    df['non_ideal'] = (df.isnull().values.any(axis=1))  | \
                               (df['Tac level (prior to am dose)'] == '<2') | \
                               (df["Tac level (prior to am dose)"].astype(str).str.contains("/"))

    # # Set boolean for non_ideal as True if all dose including and above current row is 0
    # for i in range(len(df)):
    #     if (df.loc[0:i, dose_string] == 0).all():
    #         df.loc[i, 'non_ideal'] = True

    # Create index column
    df.reset_index(inplace=True) 

    try:
        # Cum sum
        df['cum_sum'] = df.non_ideal.cumsum()
        df_agg = df[df.non_ideal==False].groupby('cum_sum').agg({'index': ['count', 'min', 'max']})
        df_agg = df_agg[df_agg['index']['count']==df_agg['index']['count'].max()]

        # Find index of data to keep
        if len(df_agg) > 1:
            print('there are multiple sets of consecutively ideal data')
            df_agg = df_agg.loc[0,:]
        min_idx = df_agg['index','min'].squeeze()
        max_idx = df_agg['index','max'].squeeze()

        # Keep longest consecutive set of ideal data
        df = df.iloc[min_idx: max_idx+1, :]
    except:
        print(f"no ideal data for patient {patient}")

    # Format patient dataframe
    df['patient'] = patient
    df = df[['Day #', 'Tac level (prior to am dose)', dose_string, 'patient']]
    df.columns = ['day', 'response', 'dose', 'patient']
    list_of_patient_df.append(df)

    return df

def cal_pred_data(df, patient, patients_to_exclude, deg):
    """
    Find calibration points and combine calibration data with the rest of data,
    to perform CURATE methods later.

    Input: 
    df - dataframe
    patient - string of patient number
    deg - degree of fitting polynomial (1 for linear, 2 for quadratic)

    Output: 
    Dataframe of calibration data and data for subsequent predictions.
    Print patients with insufficient data for calibration and <3 predictions

        """
    # Create index column
    df = df.reset_index(drop=False) 

    # Create boolean column to check if tac dose diff from next row
    df['diff_from_next'] = \
            (df['dose'] != df['dose'].shift(-1))

    # Find indexes of last rows of first 2 unique doses
    last_unique_doses_idx = [i for i, x in enumerate(df.diff_from_next) if x]

    # Create boolean column to check if tac dose diff from previous row
    df['diff_from_prev'] = \
            (df['dose'] != df['dose'].shift(1))

    # Find indexes of first rows of third unique dose
    first_unique_doses_idx = [i for i, x in enumerate(df.diff_from_prev) if x]

    # The boolean checks created useless index, diff_from_next and diff_from_prev columns,
    # drop them
    df = df.drop(['index', 'diff_from_next', 'diff_from_prev'], axis=1)

    # Combine calibration and prediction rows
    cal_pred = pd.DataFrame()

    # Do for quadratic method
    if deg == 2:
        if df['dose'].nunique() > 2:
            # If third cal point is the same as first 2 cal points, keep looking
            first_cal_dose = df['dose'][last_unique_doses_idx[0]]
            second_cal_dose = df['dose'][last_unique_doses_idx[1]]
            n = 2
            for i in range(n, len(df)+1):
                third_cal_dose = df['dose'][first_unique_doses_idx[n]]
                if (third_cal_dose == first_cal_dose) | (third_cal_dose == second_cal_dose):
                    n = n + 1

            first_cal_point = pd.DataFrame(df.iloc[last_unique_doses_idx[0],:]).T
            second_cal_point = pd.DataFrame(df.iloc[last_unique_doses_idx[1],:]).T
            third_cal_point = pd.DataFrame(df.iloc[first_unique_doses_idx[n],:]).T
            rest_of_data = df.iloc[first_unique_doses_idx[n]+1:,:]
            cal_pred = pd.concat([first_cal_point, second_cal_point, third_cal_point, 
                                        rest_of_data]).reset_index(drop=True)
        else:
            patients_to_exclude.append(str(patient))
            print(f"Patient #{patient} has insufficient unique dose-response pairs for calibration (for quad)!")
            
    # Do for linear method
    if deg == 1:
        if df['dose'].nunique() > 1:
            first_cal_point = pd.DataFrame(df.iloc[last_unique_doses_idx[0],:]).T
            second_cal_point = pd.DataFrame(df.iloc[first_unique_doses_idx[1],:]).T
            rest_of_data = df.iloc[first_unique_doses_idx[1]+1:,:]
            cal_pred = pd.concat([first_cal_point, second_cal_point, 
                                        rest_of_data]).reset_index(drop=True)
        else:
            patients_to_exclude.append(str(patient))
            print(f"Patient #{patient} has insufficient unique dose-response pairs for calibration (for linear)!")

    # Print error msg if number of predictions is less than 3
    if (len(cal_pred) - (deg + 1) < 3) and (df['dose'].nunique() > deg):
        patients_to_exclude.append(str(patient))
        if deg == 1:
            error_string = '(for linear)'
        else:
            error_string = '(for quadratic)'
            
        print(f"Patient #{patient} has insufficient/<3 predictions ({len(cal_pred) - (deg + 1)} predictions) {error_string}!")

    # Add "type" column
    cal_pred['type'] = ""
    if deg == 1:
        cal_pred['type'] = 'linear'
    else:
        cal_pred['type'] = 'quadratic'

    return cal_pred, patients_to_exclude

def keep_target_patients(patient, patients_to_exclude_linear, patients_to_exclude_quad, cal_pred_linear, cal_pred_quad, \
                        list_of_cal_pred_df):
    """
    Keep patient data with sufficient dose-response pairs and predictions
    
    Details:
    There are 4 scenarios depending on inclusion/exlusion of patient for linear/quad methods.
    Append patient data to list_of_cal_pred_df only if there's sufficient dose-response pairs and predictions.
    
    Input:
    patient
    patients_to_exclude_linear - list of patients to be excluded from linear methods
    patients_to_exclude_quad - list of patients to be excluded from quad methods
    cal_pred_linear - individual patient dataframe with calibration and efficacy-driven dosing data for linear methods
    cal_pred_quad - same as above, for quad methods
    list_of_cal_pred_df - list of all cal_pred dataframes of each patient from first to current patient in loop
    
    Output:
    cal_pred
        - individual patient dataframe of calibration and efficacy-driven dosing, 
          that has sufficient dose-response pairs and predictions
    list_of_cal_pred_df 
        - list of all cal_pred dataframes of each patient 
          from first to current patient in loop
    """
    if (patient not in patients_to_exclude_linear) and (patient not in patients_to_exclude_quad):
        cal_pred = pd.concat([cal_pred_linear, cal_pred_quad])
        list_of_cal_pred_df.append(cal_pred)
    elif patient in patients_to_exclude_linear and (patient not in patients_to_exclude_quad):
        cal_pred = cal_pred_quad
        list_of_cal_pred_df.append(cal_pred)
    elif patient not in patients_to_exclude_linear and (patient in patients_to_exclude_quad):
        cal_pred = cal_pred_linear
        list_of_cal_pred_df.append(cal_pred)
    else: cal_pred = pd.DataFrame()

    cal_pred.reset_index(inplace=True, drop=True)

    return cal_pred, list_of_cal_pred_df

# Prepare patient dataframe for prediction and apply method
def apply_methods(cal_pred, patient, patients_to_exclude_linear, patients_to_exclude_quad,
                  cal_pred_linear, cal_pred_quad, list_of_result_df, five_fold_cross_val_results_summary):
    
    """
    If cal_pred is filled, create result dataframe and apply all methods.
    
    Input:
    cal_pred - combined dataframe of cal_pred_linear and cal_pred_quad that satisfy minimum criteria.
    patient
    patients_to_exclude_linear - list of patients to exclude for linear methods
    patients_to_exclude_linear - list of patients to exclude for quadratic methods
    cal_pred_linear - individual patient dataframe with calibration and efficacy-driven dosing data for linear methods
    cal_pred_quad - same as above, for quad methods
    list_of_result_df - list of result dataframe from each patient
    
    Output:
    list_of_result_df - list of result dataframe from each patient
    """
    
    if len(cal_pred) != 0:

        # Create result dataFrame
        max_count_input = len(cal_pred_linear)

        col_names = ['patient', 'method', 'pred_day'] + \
                    ['fit_dose_' + str(i) for i in range(1, max_count_input + 1)] + \
                    ['fit_response_' + str(i) for i in range(1, max_count_input + 1)] + \
                    ['day_' + str(i) for i in range(1, max_count_input + 1)] + \
                    ['weight_' + str(i) for i in range(1, max_count_input + 1)] + \
                    ['half_life', 'dose', 'response', 'prev_coeff_2x', 'prev_coeff_1x', 'prev_coeff_0x',\
                     'prev_deviation', 'coeff_2x', 'coeff_1x', 'coeff_0x', 'prediction', 'deviation',
                     'abs_deviation']
        result = pd.DataFrame(columns=col_names)    

        if patient not in patients_to_exclude_linear:
            deg = 1

            if not isinstance(five_fold_cross_val_results_summary, pd.DataFrame):
                list_of_result_df = Cum(deg, cal_pred_linear, result, 'L_Cum_wo_origin', list_of_result_df, 'wo_origin', tau="")
                list_of_result_df = Cum(deg, cal_pred_linear, result, 'L_Cum_wo_origin_tau', list_of_result_df, 'wo_origin', tau=1)
                list_of_result_df = Cum(deg, cal_pred_linear, result, 'L_Cum_origin_dp', list_of_result_df, 'origin_dp', tau="")
                list_of_result_df = Cum(deg, cal_pred_linear, result, 'L_Cum_origin_dp_tau', list_of_result_df, 'origin_dp', tau=1)
                list_of_result_df = Cum(deg, cal_pred_linear, result, 'L_Cum_origin_int', list_of_result_df, 'origin_int', tau="")
                list_of_result_df = Cum(deg, cal_pred_linear, result, 'L_Cum_origin_int_tau', list_of_result_df, 'origin_int', tau=1)
                list_of_result_df = PPM(deg, cal_pred_linear, result, 'L_PPM_wo_origin', list_of_result_df, 'wo_origin', tau="")
                list_of_result_df = PPM(deg, cal_pred_linear, result, 'L_PPM_wo_origin_tau', list_of_result_df, 'wo_origin', tau=1)
                list_of_result_df = PPM(deg, cal_pred_linear, result, 'L_PPM_origin_dp', list_of_result_df, 'origin_dp', tau="")
                list_of_result_df = PPM(deg, cal_pred_linear, result, 'L_PPM_origin_dp_tau', list_of_result_df, 'origin_dp', tau=1)
                list_of_result_df = PPM(deg, cal_pred_linear, result, 'L_PPM_origin_int', list_of_result_df, 'origin_int', tau="")
                list_of_result_df = PPM(deg, cal_pred_linear, result, 'L_PPM_origin_int_tau', list_of_result_df, 'origin_int', tau=1)
                list_of_result_df = RW(deg, cal_pred_linear, result, 'L_RW_wo_origin', list_of_result_df, 'wo_origin', tau="")
                list_of_result_df = RW(deg, cal_pred_linear, result, 'L_RW_wo_origin_tau', list_of_result_df, 'wo_origin', tau=1)
                list_of_result_df = RW(deg, cal_pred_linear, result, 'L_RW_origin_dp', list_of_result_df, 'origin_dp', tau="")
                list_of_result_df = RW(deg, cal_pred_linear, result, 'L_RW_origin_dp_tau', list_of_result_df, 'origin_dp', tau=1)
                list_of_result_df = RW(deg, cal_pred_linear, result, 'L_RW_origin_int', list_of_result_df, 'origin_int', tau="")
                list_of_result_df = RW(deg, cal_pred_linear, result, 'L_RW_origin_int_tau', list_of_result_df, 'origin_int', tau=1)

            # Pop tau
            else:
                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'L_Cum_wo_origin_tau'], 'pop_half_life'])
                list_of_result_df = Cum(deg, cal_pred_linear, result, 'L_Cum_wo_origin_pop_tau', list_of_result_df, \
                    'wo_origin', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'L_Cum_origin_dp_tau'], 'pop_half_life'])
                list_of_result_df = Cum(deg, cal_pred_linear, result, 'L_Cum_origin_dp_pop_tau', list_of_result_df, \
                    'origin_dp', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'L_Cum_origin_int_tau'], 'pop_half_life'])
                list_of_result_df = Cum(deg, cal_pred_linear, result, 'L_Cum_origin_int_pop_tau', list_of_result_df, \
                    'origin_int', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'L_PPM_wo_origin_tau'], 'pop_half_life'])
                list_of_result_df = PPM(deg, cal_pred_linear, result, 'L_PPM_wo_origin_pop_tau', list_of_result_df, \
                    'wo_origin', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'L_PPM_origin_dp_tau'], 'pop_half_life'])
                list_of_result_df = PPM(deg, cal_pred_linear, result, 'L_PPM_origin_dp_pop_tau', list_of_result_df, \
                    'origin_dp', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'L_PPM_origin_int_tau'], 'pop_half_life'])
                list_of_result_df = PPM(deg, cal_pred_linear, result, 'L_PPM_origin_int_pop_tau', list_of_result_df, \
                    'origin_int', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'L_RW_wo_origin_tau'], 'pop_half_life'])
                list_of_result_df = RW(deg, cal_pred_linear, result, 'L_RW_wo_origin_pop_tau', list_of_result_df, \
                    'wo_origin', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'L_RW_origin_dp_tau'], 'pop_half_life'])
                list_of_result_df = RW(deg, cal_pred_linear, result, 'L_RW_origin_dp_pop_tau', list_of_result_df, \
                    'origin_dp', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'L_RW_origin_int_tau'], 'pop_half_life'])
                list_of_result_df = RW(deg, cal_pred_linear, result, 'L_RW_origin_int_pop_tau', list_of_result_df, \
                    'origin_int', tau=1, half_life=[pop_half_life])

        if patient not in patients_to_exclude_quad:
            deg = 2

            if not isinstance(five_fold_cross_val_results_summary, pd.DataFrame):
                list_of_result_df = Cum(deg, cal_pred_quad, result, 'Q_Cum_wo_origin', list_of_result_df, 'wo_origin', tau="")
                list_of_result_df = Cum(deg, cal_pred_quad, result, 'Q_Cum_wo_origin_tau', list_of_result_df, 'wo_origin', tau=1)
                list_of_result_df = Cum(deg, cal_pred_quad, result, 'Q_Cum_origin_dp', list_of_result_df, 'origin_dp', tau="")
                list_of_result_df = Cum(deg, cal_pred_quad, result, 'Q_Cum_origin_dp_tau', list_of_result_df, 'origin_dp', tau=1)
                list_of_result_df = Cum(deg, cal_pred_quad, result, 'Q_Cum_origin_int', list_of_result_df, 'origin_int', tau="")
                list_of_result_df = Cum(deg, cal_pred_quad, result, 'Q_Cum_origin_int_tau', list_of_result_df, 'origin_int', tau=1)
                list_of_result_df = PPM(deg, cal_pred_quad, result, 'Q_PPM_wo_origin', list_of_result_df, 'wo_origin', tau="")
                list_of_result_df = PPM(deg, cal_pred_quad, result, 'Q_PPM_wo_origin_tau', list_of_result_df, 'wo_origin', tau=1)
                list_of_result_df = PPM(deg, cal_pred_quad, result, 'Q_PPM_origin_dp', list_of_result_df, 'origin_dp', tau="")
                list_of_result_df = PPM(deg, cal_pred_quad, result, 'Q_PPM_origin_dp_tau', list_of_result_df, 'origin_dp', tau=1)
                list_of_result_df = PPM(deg, cal_pred_quad, result, 'Q_PPM_origin_int', list_of_result_df, 'origin_int', tau="")
                list_of_result_df = PPM(deg, cal_pred_quad, result, 'Q_PPM_origin_int_tau', list_of_result_df, 'origin_int', tau=1)
                list_of_result_df = RW(deg, cal_pred_quad, result, 'Q_RW_wo_origin', list_of_result_df, 'wo_origin', tau="")
                list_of_result_df = RW(deg, cal_pred_quad, result, 'Q_RW_wo_origin_tau', list_of_result_df, 'wo_origin', tau=1)
                list_of_result_df = RW(deg, cal_pred_quad, result, 'Q_RW_origin_dp', list_of_result_df, 'origin_dp', tau="")
                list_of_result_df = RW(deg, cal_pred_quad, result, 'Q_RW_origin_dp_tau', list_of_result_df, 'origin_dp', tau=1)
                list_of_result_df = RW(deg, cal_pred_quad, result, 'Q_RW_origin_int', list_of_result_df, 'origin_int', tau="")
                list_of_result_df = RW(deg, cal_pred_quad, result, 'Q_RW_origin_int_tau', list_of_result_df, 'origin_int', tau=1)

            # Pop tau
            else:
                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'Q_Cum_wo_origin_tau'], 'pop_half_life'])
                list_of_result_df = Cum(deg, cal_pred_quad, result, 'Q_Cum_wo_origin_pop_tau', list_of_result_df, \
                    'wo_origin', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'Q_Cum_origin_dp_tau'], 'pop_half_life'])
                list_of_result_df = Cum(deg, cal_pred_quad, result, 'Q_Cum_origin_dp_pop_tau', list_of_result_df, \
                    'origin_dp', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'Q_Cum_origin_int_tau'], 'pop_half_life'])
                list_of_result_df = Cum(deg, cal_pred_quad, result, 'Q_Cum_origin_int_pop_tau', list_of_result_df, \
                    'origin_int', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'Q_PPM_wo_origin_tau'], 'pop_half_life'])
                list_of_result_df = PPM(deg, cal_pred_quad, result, 'Q_PPM_wo_origin_pop_tau', list_of_result_df, \
                    'wo_origin', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'Q_PPM_origin_dp_tau'], 'pop_half_life'])
                list_of_result_df = PPM(deg, cal_pred_quad, result, 'Q_PPM_origin_dp_pop_tau', list_of_result_df, \
                    'origin_dp', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'Q_PPM_origin_int_tau'], 'pop_half_life'])
                list_of_result_df = PPM(deg, cal_pred_quad, result, 'Q_PPM_origin_int_pop_tau', list_of_result_df, \
                    'origin_int', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'Q_RW_wo_origin_tau'], 'pop_half_life'])
                list_of_result_df = RW(deg, cal_pred_quad, result, 'Q_RW_wo_origin_pop_tau', list_of_result_df, \
                    'wo_origin', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'Q_RW_origin_dp_tau'], 'pop_half_life'])
                list_of_result_df = RW(deg, cal_pred_quad, result, 'Q_RW_origin_dp_pop_tau', list_of_result_df, \
                    'origin_dp', tau=1, half_life=[pop_half_life])

                pop_half_life = float(five_fold_cross_val_results_summary.loc[five_fold_cross_val_results_summary.index\
                    [five_fold_cross_val_results_summary.method == 'Q_RW_origin_int_tau'], 'pop_half_life'])
                list_of_result_df = RW(deg, cal_pred_quad, result, 'Q_RW_origin_int_pop_tau', list_of_result_df, \
                    'origin_int', tau=1, half_life=[pop_half_life])

    return list_of_result_df

def Cum(deg, cal_pred, result, method_string, list_of_result_df, origin_inclusion='wo_origin', tau="", half_life=np.arange(3.5, 41.5, 1)):
    """
    Execute cumulative approach for each variation of origin inclusion, type, and tau inclusion. 

    Details:
    (If the chosen method has tau, the following details are looped for each variation of half_life)
    Add a fresh line to result dataframe for each prediction and fill in the line with patient, method,
    prediction day (and half-life). Fill in the dose-response pairs and days of the pairs that are 
    required to fit the model, according to the origin inclusion chosen. Fill in the dose-response 
    pair of the day of the prediction to calculate the deviation of the prediction. (If chosen method 
    has tau, calculate the weights based on the number of days from prediction day). Fit the model 
    (with weight if chosen method has tau). Calculate the prediction, deviation, and absolute deviation, 
    and fill in the results to the result dataframe.

    Input:
    deg - degree of polynomial fit, 1 for linear, 2 for quadratic
    cal_pred - dataframe of calibration and efficacy-driven dosing data, 
                cal_pred_linear for linear, cal_pred_quad for quad
    result - dataframe of results
    method_string - string of method
    list_of_result_df - list of result dataframe from each patient

    Output:
    list_of_result_df
    """
    j = 0

    result = result[0:0]
    
    # half_life = np.arange(3.5, 41.5, 1)

    if tau == 1:

        # Loop through all half_lives
        for k in range(len(half_life)):

            result = result[0:0]

            # Loop through each prediction
            for i in range(deg + 1, len(cal_pred)):
            
                result.loc[j, 'patient'] = cal_pred.loc[i, 'patient']
                result.loc[j, 'method'] = method_string
                result.loc[j, 'pred_day'] = cal_pred.loc[i, 'day']
                if (origin_inclusion == 'wo_origin') or (origin_inclusion =='origin_int'):
                    result.loc[j, 'fit_dose_1':'fit_dose_' + str(i)] = cal_pred.loc[0:i-1, 'dose'].to_numpy()
                    result.loc[j, 'fit_response_1':'fit_response_' + str(i)] = cal_pred.loc[0:i-1, 'response'].to_numpy()
                    result.loc[j, 'day_1':'day_' + str(i)] = cal_pred.loc[0:i-1, 'day'].to_numpy()
                elif origin_inclusion == 'origin_dp':
                    result.loc[j, 'fit_dose_1'] = 0
                    result.loc[j, 'fit_dose_2':'fit_dose_' + str(i+1)] = cal_pred.loc[0:i-1, 'dose'].to_numpy()
                    result.loc[j, 'fit_response_1'] = 0
                    result.loc[j, 'fit_response_2':'fit_response_' + str(i+1)] = cal_pred.loc[0:i-1, 'response'].to_numpy()
                    result.loc[j, 'day_1'] = cal_pred.loc[0, 'day'] - 1
                    result.loc[j, 'day_2':'day_' + str(i+1)] = cal_pred.loc[0:i-1, 'day'].to_numpy()
                
                result.loc[j, 'dose'] = cal_pred.loc[i, 'dose']
                result.loc[j, 'response'] = cal_pred.loc[i, 'response']
                result.loc[j, 'half_life'] = half_life[k]

                # Calculate and fill in weight
                fixed_pred_day = result.loc[j, 'pred_day']
                fixed_half_life = result.loc[j, 'half_life']
                if (origin_inclusion == 'wo_origin') or (origin_inclusion =='origin_int'):
                    day_array = result.loc[j, 'day_1':'day_' + str(i)].astype(float)

                    # Fill in weight array
                    result.loc[j, 'weight_1':'weight_' + str(i)] = np.exp(-(24*(fixed_pred_day - day_array)*(math.log(2)/fixed_half_life))).to_numpy()

                elif origin_inclusion == 'origin_dp':
                    day_array = result.loc[j, 'day_1':'day_' + str(i+1)].astype(float)

                    # Fill in weight array
                    result.loc[j, 'weight_1':'weight_' + str(i+1)] = np.exp(-(24*(fixed_pred_day - day_array)*(math.log(2)/fixed_half_life))).to_numpy()

                # Curve fit equation
                if origin_inclusion == 'wo_origin' or origin_inclusion == 'origin_int':
                    X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(i)].to_numpy()
                    y = result.loc[j, 'fit_response_1':'fit_response_' + str(i)].to_numpy()
                    weight = result.loc[j, 'weight_1':'weight_' + str(i)]
                elif origin_inclusion == 'origin_dp':
                    X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(i+1)].to_numpy()
                    y = result.loc[j, 'fit_response_1':'fit_response_' + str(i+1)].to_numpy()
                    weight = result.loc[j, 'weight_1':'weight_' + str(i+1)]
                
                if deg == 1:
                    if origin_inclusion == 'origin_int':
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=True).fit(X, y, weight)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                    else:
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=False).fit(X, y, weight)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                else:
                    if origin_inclusion == 'origin_int':
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=True).fit(X, y, weight)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                        result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                    else:
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=False).fit(X, y, weight)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                        result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]

                # Calculate prediction and deviation
                
                prediction = fitted_model.predict(poly_reg.fit_transform([[cal_pred.loc[i, 'dose']]]))[0]

                result.loc[j, 'prediction'] = prediction
                deviation = cal_pred.loc[i, 'response'] - prediction
                result.loc[j, 'deviation'] = deviation
                abs_deviation = abs(deviation)
                result.loc[j, 'abs_deviation'] = abs_deviation
                
                j = j + 1

            list_of_result_df.append(result)

    else:

        result = result[0:0]

        # Loop through each prediction
        for i in range(deg + 1, len(cal_pred)):
    
            result.loc[j, 'patient'] = cal_pred.loc[i, 'patient']
            result.loc[j, 'method'] = method_string
            result.loc[j, 'pred_day'] = cal_pred.loc[i, 'day']
            if (origin_inclusion == 'wo_origin') or (origin_inclusion =='origin_int'):
                result.loc[j, 'fit_dose_1':'fit_dose_' + str(i)] = cal_pred.loc[0:i-1, 'dose'].to_numpy()
                result.loc[j, 'fit_response_1':'fit_response_' + str(i)] = cal_pred.loc[0:i-1, 'response'].to_numpy()
                result.loc[j, 'day_1':'day_' + str(i)] = cal_pred.loc[0:i-1, 'day'].to_numpy()
            elif origin_inclusion == 'origin_dp':
                result.loc[j, 'fit_dose_1'] = 0
                result.loc[j, 'fit_dose_2':'fit_dose_' + str(i+1)] = cal_pred.loc[0:i-1, 'dose'].to_numpy()
                result.loc[j, 'fit_response_1'] = 0
                result.loc[j, 'fit_response_2':'fit_response_' + str(i+1)] = cal_pred.loc[0:i-1, 'response'].to_numpy()
                result.loc[j, 'day_1'] = cal_pred.loc[0, 'day'] - 1
                result.loc[j, 'day_2':'day_' + str(i+1)] = cal_pred.loc[0:i-1, 'day'].to_numpy()
            
            result.loc[j, 'dose'] = cal_pred.loc[i, 'dose']
            result.loc[j, 'response'] = cal_pred.loc[i, 'response']

            # Curve fit equation
            if origin_inclusion == 'wo_origin' or origin_inclusion == 'origin_int':
                X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(i)].to_numpy()
                y = result.loc[j, 'fit_response_1':'fit_response_' + str(i)].to_numpy()
            elif origin_inclusion == 'origin_dp':
                X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(i+1)].to_numpy()
                y = result.loc[j, 'fit_response_1':'fit_response_' + str(i+1)].to_numpy()
            
            if deg == 1:
                if origin_inclusion == 'origin_int':
                    poly_reg = PolynomialFeatures(degree=deg)
                    X = X.reshape(-1,1)
                    X = poly_reg.fit_transform(X)
                    fitted_model = LinearRegression(fit_intercept=True).fit(X, y)
                    result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                    result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                else:
                    poly_reg = PolynomialFeatures(degree=deg)
                    X = X.reshape(-1,1)
                    X = poly_reg.fit_transform(X)
                    fitted_model = LinearRegression(fit_intercept=False).fit(X, y)
                    result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                    result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
            else:
                if origin_inclusion == 'origin_int':
                    poly_reg = PolynomialFeatures(degree=deg)
                    X = X.reshape(-1,1)
                    X = poly_reg.fit_transform(X)
                    fitted_model = LinearRegression(fit_intercept=True).fit(X, y)
                    result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                    result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                    result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                else:
                    poly_reg = PolynomialFeatures(degree=deg)
                    X = X.reshape(-1,1)
                    X = poly_reg.fit_transform(X)
                    fitted_model = LinearRegression(fit_intercept=False).fit(X, y)
                    result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                    result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                    result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                
            # Calculate prediction and deviation
            prediction = fitted_model.predict(poly_reg.fit_transform([[cal_pred.loc[i, 'dose']]]))[0]

            result.loc[j, 'prediction'] = prediction
            deviation = cal_pred.loc[i, 'response'] - prediction
            result.loc[j, 'deviation'] = deviation
            abs_deviation = abs(deviation)
            result.loc[j, 'abs_deviation'] = abs_deviation
            
            j = j + 1

        list_of_result_df.append(result)

    return list_of_result_df

def PPM(deg, cal_pred, result, method_string, list_of_result_df, origin_inclusion='wo_origin', tau="", half_life=np.arange(3.5, 41.5, 1)):
    """
    Execute cumulative approach for each variation of origin inclusion, type, and tau inclusion. 

    Details:
    (If the chosen method has tau, the following details are looped for each variation of half_life)
    Add a fresh line to result dataframe for each prediction and fill in the line with patient, method,
    prediction day (and half-life). For first prediction, fill up dose and response to be fitted. 
    Then fill in coefficients, prediction and deviation of fit. For second and following predictions, 
    fill up dose and response of prediction day, fill in previous coefficients and deviations,
    shift last coefficient by previous deviation, fill in prediction and deviation of fit.

    Input:
    deg - degree of polynomial fit, 1 for linear, 2 for quadratic
    cal_pred - dataframe of calibration and efficacy-driven dosing data, 
                cal_pred_linear for linear, cal_pred_quad for quad
    result - dataframe of results
    method_string - string of method
    list_of_result_df - list of result dataframe from each patient

    Output:
    list_of_result_df
    """
    j = 0 # Counter for number of rows in "result"
        
    # half_life = np.arange(3.5, 41.5, 1)

    if tau == 1:

        # Loop through all half_lives
        for k in range(len(half_life)):

            result = result[0:0]

            j = 0
                
            # Loop through each prediction
            for i in range(deg + 1, len(cal_pred)):

                result.loc[j, 'half_life'] = half_life[k]

                # First prediction
                if j == 0:                    
                    result.loc[j, 'patient'] = cal_pred.loc[i, 'patient']
                    result.loc[j, 'method'] = method_string
                    result.loc[j, 'pred_day'] = cal_pred.loc[i, 'day']

                    if (origin_inclusion == 'wo_origin') or (origin_inclusion == 'origin_int'):
                        result.loc[j, 'fit_dose_1':'fit_dose_' + str(i)] = cal_pred.loc[0:i-1, 'dose'].to_numpy()
                        result.loc[j, 'fit_response_1':'fit_response_' + str(i)] = cal_pred.loc[0:i-1, 'response'].to_numpy()
                        result.loc[j, 'day_1':'day_' + str(i)] = cal_pred.loc[0:i-1, 'day'].to_numpy()
                    elif origin_inclusion == 'origin_dp':
                        result.loc[j, 'fit_dose_1'] = 0
                        result.loc[j, 'fit_dose_2':'fit_dose_' + str(i+1)] = cal_pred.loc[0:i-1, 'dose'].to_numpy()
                        result.loc[j, 'fit_response_1'] = 0
                        result.loc[j, 'fit_response_2':'fit_response_' + str(i+1)] = cal_pred.loc[0:i-1, 'response'].to_numpy()
                        result.loc[j, 'day_1'] = cal_pred.loc[0, 'day'] - 1
                        result.loc[j, 'day_2':'day_' + str(i+1)] = cal_pred.loc[0:i-1, 'day'].to_numpy()

                    result.loc[j, 'dose'] = cal_pred.loc[i, 'dose']
                    result.loc[j, 'response'] = cal_pred.loc[i, 'response']
                    
                    # Calculate and fill in weight
                    fixed_pred_day = result.loc[j, 'pred_day']
                    fixed_half_life = result.loc[j, 'half_life']
                    if (origin_inclusion == 'wo_origin') or (origin_inclusion =='origin_int'):
                        day_array = result.loc[j, 'day_1':'day_' + str(i)].astype(float)

                        # Fill in weight array
                        result.loc[j, 'weight_1':'weight_' + str(i)] = np.exp(-(24*(fixed_pred_day - day_array)*(math.log(2)/fixed_half_life))).to_numpy()

                    elif origin_inclusion == 'origin_dp':
                        day_array = result.loc[j, 'day_1':'day_' + str(i+1)].astype(float)

                        # Fill in weight array
                        result.loc[j, 'weight_1':'weight_' + str(i+1)] = np.exp(-(24*(fixed_pred_day - day_array)*(math.log(2)/fixed_half_life))).to_numpy()

                    # Curve fit equation
                    if origin_inclusion == 'wo_origin' or origin_inclusion == 'origin_int':
                        X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(i)].to_numpy()
                        y = result.loc[j, 'fit_response_1':'fit_response_' + str(i)].to_numpy()
                        weight = result.loc[j, 'weight_1':'weight_' + str(i)]
                    elif origin_inclusion == 'origin_dp':
                        X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(i+1)].to_numpy()
                        y = result.loc[j, 'fit_response_1':'fit_response_' + str(i+1)].to_numpy()
                        weight = result.loc[j, 'weight_1':'weight_' + str(i+1)]
                    
                    if deg == 1:
                        if origin_inclusion == 'origin_int':
                            poly_reg = PolynomialFeatures(degree=deg)
                            X = X.reshape(-1,1)
                            X = poly_reg.fit_transform(X)
                            fitted_model = LinearRegression(fit_intercept=True).fit(X, y, weight)
                            result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                            result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                        else:
                            poly_reg = PolynomialFeatures(degree=deg)
                            X = X.reshape(-1,1)
                            X = poly_reg.fit_transform(X)
                            fitted_model = LinearRegression(fit_intercept=False).fit(X, y, weight)
                            result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                            result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                    else:
                        if origin_inclusion == 'origin_int':
                            poly_reg = PolynomialFeatures(degree=deg)
                            X = X.reshape(-1,1)
                            X = poly_reg.fit_transform(X)
                            fitted_model = LinearRegression(fit_intercept=True).fit(X, y, weight)
                            result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                            result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                            result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                        else:
                            poly_reg = PolynomialFeatures(degree=deg)
                            X = X.reshape(-1,1)
                            X = poly_reg.fit_transform(X)
                            fitted_model = LinearRegression(fit_intercept=False).fit(X, y, weight)
                            result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                            result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                            result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                          
                    # Calculate prediction and deviation
                    prediction = fitted_model.predict(poly_reg.fit_transform([[cal_pred.loc[i, 'dose']]]))[0]

                    result.loc[j, 'prediction'] = prediction
                    deviation = cal_pred.loc[i, 'response'] - prediction
                    result.loc[j, 'deviation'] = deviation
                    abs_deviation = abs(deviation)
                    result.loc[j, 'abs_deviation'] = abs_deviation

                # Second prediction onwards
                else:               
                    result.loc[j, 'patient'] = cal_pred.loc[i, 'patient']
                    result.loc[j, 'method'] = method_string
                    result.loc[j, 'pred_day'] = cal_pred.loc[i, 'day']
                    result.loc[j, 'dose'] = cal_pred.loc[i, 'dose']
                    result.loc[j, 'response'] = cal_pred.loc[i, 'response']

                    # Fill in previous coeff and deviation
                    result.loc[j, 'prev_coeff_2x'] = result.loc[j-1, 'coeff_2x']
                    result.loc[j, 'prev_coeff_1x'] = result.loc[j-1, 'coeff_1x']
                    result.loc[j, 'prev_coeff_0x'] = result.loc[j-1, 'coeff_0x']
                    result.loc[j, 'prev_deviation'] = result.loc[j-1, 'deviation']

                    # Shift last coefficient by previous deviation
                    result.loc[j, 'coeff_2x'] = result.loc[j, 'prev_coeff_2x']
                    result.loc[j, 'coeff_1x'] = result.loc[j, 'prev_coeff_1x']
                    result.loc[j, 'coeff_0x'] = result.loc[j, 'prev_coeff_0x'] + result.loc[j, 'prev_deviation']

                    # Calculate prediction and deviation
                    if deg == 1:
                        prediction = result.loc[j, 'dose'] * result.loc[j, 'coeff_1x'] +  result.loc[j, 'coeff_0x']
                    else:
                        prediction = (result.loc[j, 'dose'] ** 2) * result.loc[j, 'coeff_2x'] + result.loc[j, 'dose'] * result.loc[j, 'coeff_1x'] +  result.loc[j, 'coeff_0x']
                    
                    result.loc[j, 'prediction'] = prediction
                    deviation =  result.loc[j, 'response'] - prediction
                    result.loc[j, 'deviation'] = deviation
                    abs_deviation = abs(deviation)
                    result.loc[j, 'abs_deviation'] = abs_deviation

                j = j + 1

            list_of_result_df.append(result)

    else:
        
        result = result[0:0]

        j = 0

        for i in range(deg + 1, len(cal_pred)):

            # First prediction
            if j == 0:                    
                result.loc[j, 'patient'] = cal_pred.loc[i, 'patient']
                result.loc[j, 'method'] = method_string
                result.loc[j, 'pred_day'] = cal_pred.loc[i, 'day']

                if (origin_inclusion == 'wo_origin') or (origin_inclusion == 'origin_int'):
                    result.loc[j, 'fit_dose_1':'fit_dose_' + str(i)] = cal_pred.loc[0:i-1, 'dose'].to_numpy()
                    result.loc[j, 'fit_response_1':'fit_response_' + str(i)] = cal_pred.loc[0:i-1, 'response'].to_numpy()
                    result.loc[j, 'day_1':'day_' + str(i)] = cal_pred.loc[0:i-1, 'day'].to_numpy()
                elif origin_inclusion == 'origin_dp':
                    result.loc[j, 'fit_dose_1'] = 0
                    result.loc[j, 'fit_dose_2':'fit_dose_' + str(i+1)] = cal_pred.loc[0:i-1, 'dose'].to_numpy()
                    result.loc[j, 'fit_response_1'] = 0
                    result.loc[j, 'fit_response_2':'fit_response_' + str(i+1)] = cal_pred.loc[0:i-1, 'response'].to_numpy()
                    result.loc[j, 'day_1'] = cal_pred.loc[0, 'day'] - 1
                    result.loc[j, 'day_2':'day_' + str(i+1)] = cal_pred.loc[0:i-1, 'day'].to_numpy()

                result.loc[j, 'dose'] = cal_pred.loc[i, 'dose']
                result.loc[j, 'response'] = cal_pred.loc[i, 'response']

                # Curve fit equation
                if origin_inclusion == 'wo_origin' or origin_inclusion == 'origin_int':
                    X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(deg+1)].to_numpy()
                    y = result.loc[j, 'fit_response_1':'fit_response_' + str(deg+1)].to_numpy()
                elif origin_inclusion == 'origin_dp':
                    X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(deg+2)].to_numpy()
                    y = result.loc[j, 'fit_response_1':'fit_response_' + str(deg+2)].to_numpy()
                
                if deg == 1:
                    if origin_inclusion == 'origin_int':
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=True).fit(X, y)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                    else:
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=False).fit(X, y)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                else:
                    if origin_inclusion == 'origin_int':
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=True).fit(X, y)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                        result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                    else:
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=False).fit(X, y)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                        result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                    
                # Calculate prediction and deviation
                prediction = fitted_model.predict(poly_reg.fit_transform([[cal_pred.loc[i, 'dose']]]))[0]

                result.loc[j, 'prediction'] = prediction
                deviation = cal_pred.loc[i, 'response'] - prediction
                result.loc[j, 'deviation'] = deviation
                abs_deviation = abs(deviation)
                result.loc[j, 'abs_deviation'] = abs_deviation

            # Second prediction onwards
            else:               
                result.loc[j, 'patient'] = cal_pred.loc[i, 'patient']
                result.loc[j, 'method'] = method_string
                result.loc[j, 'pred_day'] = cal_pred.loc[i, 'day']
                result.loc[j, 'dose'] = cal_pred.loc[i, 'dose']
                result.loc[j, 'response'] = cal_pred.loc[i, 'response']

                # Fill in previous coeff and deviation
                result.loc[j, 'prev_coeff_2x'] = result.loc[j-1, 'coeff_2x']
                result.loc[j, 'prev_coeff_1x'] = result.loc[j-1, 'coeff_1x']
                result.loc[j, 'prev_coeff_0x'] = result.loc[j-1, 'coeff_0x']
                result.loc[j, 'prev_deviation'] = result.loc[j-1, 'deviation']

                # Shift last coefficient by previous deviation
                result.loc[j, 'coeff_2x'] = result.loc[j, 'prev_coeff_2x']
                result.loc[j, 'coeff_1x'] = result.loc[j, 'prev_coeff_1x']
                result.loc[j, 'coeff_0x'] = result.loc[j, 'prev_coeff_0x'] + result.loc[j, 'prev_deviation']

                # Calculate prediction and deviation
                if deg == 1:
                    prediction = result.loc[j, 'dose'] * result.loc[j, 'coeff_1x'] +  result.loc[j, 'coeff_0x']
                else:
                    prediction = (result.loc[j, 'dose'] ** 2) * result.loc[j, 'coeff_2x'] + result.loc[j, 'dose'] * result.loc[j, 'coeff_1x'] +  result.loc[j, 'coeff_0x']
                
                result.loc[j, 'prediction'] = prediction
                deviation =  result.loc[j, 'response'] - prediction
                result.loc[j, 'deviation'] = deviation
                abs_deviation = abs(deviation)
                result.loc[j, 'abs_deviation'] = abs_deviation

            j = j + 1

        list_of_result_df.append(result)
        
    return list_of_result_df

def RW(deg, cal_pred, result, method_string, list_of_result_df, origin_inclusion='wo_origin', tau="", half_life=np.arange(3.5, 41.5, 1)):
    """
    Execute rolling window approach for each variation of origin inclusion, type, and tau inclusion. 

    Details:
    (If the chosen method has tau, the following details are looped for each variation of half_life)
    Add a fresh line to result dataframe for each prediction and fill in the line with patient, method,
    prediction day (and half-life). Fill in the dose-response pairs and days of the pairs that are 
    required to fit the model, based on last deg + 1 unique dose-response pairs, according to the origin 
    inclusion chosen. Fill in the dose-response pair of the day of the prediction to calculate the 
    deviation of the prediction. (If chosen method has tau, calculate the weights based on the number 
    of days from prediction day). Fit the model (with weight if chosen method has tau). Calculate the 
    prediction, deviation, and absolute deviation, and fill in the results to the result dataframe.
    
    Input:
    deg - degree of polynomial fit, 1 for linear, 2 for quadratic
    cal_pred - dataframe of calibration and efficacy-driven dosing data, 
                cal_pred_linear for linear, cal_pred_quad for quad
    result - dataframe of results
    method_string - string of method
    list_of_result_df - list of result dataframe from each patient

    Output:
    list_of_result_df
    """
    values = []
    indices = []
    
    j = 0

    result = result[0:0]

    # half_life = np.arange(3.5, 41.5, 1)

    if tau == 1:

        # Loop through all half_lives
        for k in range(len(half_life)):

            result = result[0:0]

            values = []
            indices = []

            j = 0

            # Loop through all rows and fill in result if enough unique dose-response pairs
            for i in range(0, len(cal_pred)-1):

                # Define new dose
                new_dose = cal_pred.loc[i, 'dose']

                # Add values and indices of first three unique doses into list
                if len(values) < deg + 1:

                    # Check if new dose in values list
                    # If not, append new dose into value list and append index into indices list
                    if new_dose not in values:
                        values.append(new_dose)
                        indices.append(i)

                else:
                    # Check if new dose in value list
                    # If yes: remove value and index of repeat, and append new dose and index to list.
                    if new_dose in values:
                        repeated_dose_index = values.index(new_dose)
                        values.pop(repeated_dose_index)
                        indices.pop(repeated_dose_index)

                        values.append(new_dose)
                        indices.append(i)

                    # If no: remove the first element of list and append value and index of new dose
                    else:
                        values.pop(0)
                        indices.pop(0)

                        values.append(new_dose)
                        indices.append(i)

                # Prepare dataframe of dose-response pairs for prediction
                counter_for_origin_int = 1
                if len(indices) == deg + 1:            
                    
                    result.loc[j, 'patient'] = cal_pred.loc[i + 1, 'patient']
                    result.loc[j, 'method'] = method_string
                    result.loc[j, 'pred_day'] = cal_pred.loc[i + 1, 'day']
                    result.loc[j, 'half_life'] = half_life[k]
                    
                    # Fill in dose, response, day to fit model
                    cal_pred_string = ['dose', 'response', 'day']
                    result_string = ['fit_dose', 'fit_response', 'day']
                    for string_num in range(0,len(cal_pred_string)):
                        if (origin_inclusion == 'wo_origin') or (origin_inclusion == 'origin_int'):
                            result.loc[j, result_string[string_num] + '_1'] = cal_pred.loc[indices[0], cal_pred_string[string_num]]
                            result.loc[j, result_string[string_num] + '_2'] = cal_pred.loc[indices[1], cal_pred_string[string_num]]
                            if deg == 2:
                                result.loc[j, result_string[string_num] + '_3'] = cal_pred.loc[indices[2], cal_pred_string[string_num]]
                        elif origin_inclusion == 'origin_dp':
                            result.loc[j, result_string[string_num] + '_1'] = 0
                            result.loc[j, 'day_1'] = cal_pred.loc[indices[0], 'day'] - 1
                            result.loc[j, result_string[string_num] + '_2'] = cal_pred.loc[indices[0], cal_pred_string[string_num]]
                            result.loc[j, result_string[string_num] + '_3'] = cal_pred.loc[indices[1], cal_pred_string[string_num]]
                            if deg == 2:
                                result.loc[j, result_string[string_num] + '_4'] = cal_pred.loc[indices[2], cal_pred_string[string_num]]
                    
                    # Fill in new dose and response of prediction day
                    result.loc[j, 'dose'] = cal_pred.loc[i + 1, 'dose']
                    result.loc[j, 'response'] = cal_pred.loc[i + 1, 'response']

                    # Calculate and fill in weight
                    fixed_pred_day = result.loc[j, 'pred_day']
                    fixed_half_life = result.loc[j, 'half_life']
                    if (origin_inclusion == 'wo_origin') or (origin_inclusion =='origin_int'):
                        day_array = result.loc[j, 'day_1':'day_' + str(deg+1)].astype(float)

                        # Fill in weight array
                        result.loc[j, 'weight_1':'weight_' + str(deg+1)] = np.exp(-(24*(fixed_pred_day - day_array)*(math.log(2)/fixed_half_life))).to_numpy()

                    elif origin_inclusion == 'origin_dp':
                        day_array = result.loc[j, 'day_1':'day_' + str(deg+2)].astype(float)

                        # Fill in weight array
                        result.loc[j, 'weight_1':'weight_' + str(deg+2)] = np.exp(-(24*(fixed_pred_day - day_array)*(math.log(2)/fixed_half_life))).to_numpy()

                    # Curve fit equation
                    if origin_inclusion == 'wo_origin' or origin_inclusion == 'origin_int':
                        X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(deg+1)].to_numpy()
                        y = result.loc[j, 'fit_response_1':'fit_response_' + str(deg+1)].to_numpy()
                        weight = result.loc[j, 'weight_1':'weight_' + str(deg+1)]
                    elif origin_inclusion == 'origin_dp':
                        X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(deg+2)].to_numpy()
                        y = result.loc[j, 'fit_response_1':'fit_response_' + str(deg+2)].to_numpy()
                        weight = result.loc[j, 'weight_1':'weight_' + str(deg+2)]
                    
                    if deg == 1:
                        if origin_inclusion == 'origin_int':
                            poly_reg = PolynomialFeatures(degree=deg)
                            X = X.reshape(-1,1)
                            X = poly_reg.fit_transform(X)
                            fitted_model = LinearRegression(fit_intercept=True).fit(X, y, weight)
                            result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                            result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                        else:
                            poly_reg = PolynomialFeatures(degree=deg)
                            X = X.reshape(-1,1)
                            X = poly_reg.fit_transform(X)
                            fitted_model = LinearRegression(fit_intercept=False).fit(X, y, weight)
                            result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                            result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                    else:
                        if origin_inclusion == 'origin_int':
                            poly_reg = PolynomialFeatures(degree=deg)
                            X = X.reshape(-1,1)
                            X = poly_reg.fit_transform(X)
                            fitted_model = LinearRegression(fit_intercept=True).fit(X, y, weight)
                            result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                            result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                            result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                        else:
                            poly_reg = PolynomialFeatures(degree=deg)
                            X = X.reshape(-1,1)
                            X = poly_reg.fit_transform(X)
                            fitted_model = LinearRegression(fit_intercept=False).fit(X, y, weight)
                            result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                            result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                            result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                    
                    # Calculate prediction and deviation
                    prediction = fitted_model.predict(poly_reg.fit_transform([[cal_pred.loc[i, 'dose']]]))[0]
                        
                    result.loc[j, 'prediction'] = prediction
                    deviation = cal_pred.loc[i + 1, 'response'] - prediction
                    result.loc[j, 'deviation'] = deviation
                    abs_deviation = abs(deviation)
                    result.loc[j, 'abs_deviation'] = abs_deviation

                    counter_for_origin_int = counter_for_origin_int + 1
                    
                j = j + 1
            
            list_of_result_df.append(result)
        
    else:
        # Loop through rows
        for i in range(0, len(cal_pred)-1):

            # Define new dose
            new_dose = cal_pred.loc[i, 'dose']

            # Add values and indices of first three unique doses into list
            if len(values) < deg + 1:

                # Check if new dose in values list
                # If not, append new dose into value list and append index into indices list
                if new_dose not in values:
                    values.append(new_dose)
                    indices.append(i)

            else:
                # Check if new dose in value list
                # If yes: remove value and index of repeat, and append new dose and index to list.
                if new_dose in values:
                    repeated_dose_index = values.index(new_dose)
                    values.pop(repeated_dose_index)
                    indices.pop(repeated_dose_index)

                    values.append(new_dose)
                    indices.append(i)

                # If no: remove the first element of list and append value and index of new dose
                else:
                    values.pop(0)
                    indices.pop(0)

                    values.append(new_dose)
                    indices.append(i)

            # Prepare dataframe of dose-response pairs for prediction
            counter_for_origin_int = 1
            if len(indices) == deg + 1:            

                result.loc[j, 'patient'] = cal_pred.loc[i + 1, 'patient']
                result.loc[j, 'method'] = method_string
                result.loc[j, 'pred_day'] = cal_pred.loc[i + 1, 'day']
                
                # Fill in dose, response, day to fit model
                cal_pred_string = ['dose', 'response', 'day']
                result_string = ['fit_dose', 'fit_response', 'day']
                for string_num in range(0,len(cal_pred_string)):
                    if (origin_inclusion == 'wo_origin') or (origin_inclusion == 'origin_int'):
                        result.loc[j, result_string[string_num] + '_1'] = cal_pred.loc[indices[0], cal_pred_string[string_num]]
                        result.loc[j, result_string[string_num] + '_2'] = cal_pred.loc[indices[1], cal_pred_string[string_num]]
                        if deg == 2:
                            result.loc[j, result_string[string_num] + '_3'] = cal_pred.loc[indices[2], cal_pred_string[string_num]]
                    elif origin_inclusion == 'origin_dp':
                        result.loc[j, result_string[string_num] + '_1'] = 0
                        result.loc[j, 'day_1'] = cal_pred.loc[indices[0], 'day'] - 1
                        result.loc[j, result_string[string_num] + '_2'] = cal_pred.loc[indices[0], cal_pred_string[string_num]]
                        result.loc[j, result_string[string_num] + '_3'] = cal_pred.loc[indices[1], cal_pred_string[string_num]]
                        if deg == 2:
                            result.loc[j, result_string[string_num] + '_4'] = cal_pred.loc[indices[2], cal_pred_string[string_num]]
                
                # Fill in new dose and response of prediction day
                result.loc[j, 'dose'] = cal_pred.loc[i + 1, 'dose']
                result.loc[j, 'response'] = cal_pred.loc[i + 1, 'response']
                
                # Curve fit equation
                if origin_inclusion == 'wo_origin' or origin_inclusion == 'origin_int':
                    X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(deg+1)].to_numpy()
                    y = result.loc[j, 'fit_response_1':'fit_response_' + str(deg+1)].to_numpy()
                elif origin_inclusion == 'origin_dp':
                    X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(deg+2)].to_numpy()
                    y = result.loc[j, 'fit_response_1':'fit_response_' + str(deg+2)].to_numpy()
                
                if deg == 1:
                    if origin_inclusion == 'origin_int':
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=True).fit(X, y)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                    else:
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=False).fit(X, y)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                else:
                    if origin_inclusion == 'origin_int':
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=True).fit(X, y)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                        result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                    else:
                        poly_reg = PolynomialFeatures(degree=deg)
                        X = X.reshape(-1,1)
                        X = poly_reg.fit_transform(X)
                        fitted_model = LinearRegression(fit_intercept=False).fit(X, y)
                        result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
                        result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                        result.loc[j, 'coeff_2x'] = fitted_model.coef_[2]
                    
                # Calculate prediction and deviation
                prediction = fitted_model.predict(poly_reg.fit_transform([[cal_pred.loc[i, 'dose']]]))[0]

                result.loc[j, 'prediction'] = prediction
                deviation = cal_pred.loc[i + 1, 'response'] - prediction
                result.loc[j, 'deviation'] = deviation
                abs_deviation = abs(deviation)
                result.loc[j, 'abs_deviation'] = abs_deviation

                counter_for_origin_int = counter_for_origin_int + 1
                
            j = j + 1
        
        list_of_result_df.append(result)

    return list_of_result_df

# Combine dataframes of individual patients
def format_result_df(cal_pred, result_df):
    """
    Rerrange column names, make patient and prediction day series of integers, sort, reset index
    
    Input: 
    cal_pred - dataframe of calibration and efficacy-driven dosing data for all patients
    result_df - results after applying all methods to patients
    
    Output:
    result_df - Formatted result_df
    """
    max_count_input = cal_pred[cal_pred['type']=='linear'].groupby('patient').count().max()['dose']
    col_names = ['patient', 'method', 'pred_day'] + \
                ['fit_dose_' + str(i) for i in range(1, max_count_input + 1)] + \
                ['fit_response_' + str(i) for i in range(1, max_count_input + 1)] + \
                ['day_' + str(i) for i in range(1, max_count_input + 1)] + \
                ['weight_' + str(i) for i in range(1, max_count_input + 1)] + \
                ['half_life', 'dose', 'response', 'prev_coeff_2x', 'prev_coeff_1x', 'prev_coeff_0x',\
                 'prev_deviation', 'coeff_2x', 'coeff_1x', 'coeff_0x', 'prediction', 'deviation',
                 'abs_deviation']
    result_df = result_df[col_names]
    result_df.patient = result_df.patient.apply(int)
    result_df.pred_day = result_df.pred_day.apply(int)
    result_df.sort_values(['patient', 'method', 'half_life', 'pred_day'], inplace=True)
    result_df = result_df.reset_index(drop=True)
    
    return result_df

if __name__ == '__main__':
    # Get arguments
    import argparse
    parser = argparse.ArgumentParser(description='Implement CURATE models and retrieve results')
    parser.add_argument("-p", "--pop_tau", type=str, default=False)
    parser.add_argument("-d", "--dose", type=str, default='total')
    parser.add_argument("-C", "--cross_val_method", type=str, default='LOOCV')
    args = parser.parse_args()
    
    print('starting profile generation...')

    original_stdout = sys.stdout
    with open('patients_to_exclude.txt', 'w') as f:
        sys.stdout = f
        execute_CURATE(dose=args.dose)
    sys.stdout = original_stdout
    print('end of profile generation for models without pop tau')

    if args.pop_tau:
        execute_CURATE_and_update_pop_tau_results(args.dose, args.cross_val_method)
        print('end of profile generation for models with and without pop tau')