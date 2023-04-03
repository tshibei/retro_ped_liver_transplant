from openpyxl import load_workbook
from scipy.optimize import curve_fit
import math
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.pylab as pylab
import matplotlib.patches as patches
import numpy as np
from scipy import stats
import seaborn as sns
from statistics import mean
import sys
pd.options.mode.chained_assignment = None  # default='warn'
from plotting import *
from scipy import interpolate

# CURATE
def execute_CURATE():
    """ 
    Execute CURATE.

    Output: 
    'CURATE_results.xlsx':
            Excel sheet with cleaned patient dataframe, 
            dataframe for calibration and efficacy-driven dosing, 
            result of all methods.
    """
    # Generate profiles and join dataframes
    patients_to_exclude_linear, list_of_patient_df, list_of_cal_pred_df, list_of_result_df = generate_profiles()
    df, cal_pred, result_df = join_dataframes(list_of_patient_df, list_of_cal_pred_df, list_of_result_df)

    # Print patients to exclude anad ouput dataframes to excel as individual sheets
    print_patients_to_exclude(patients_to_exclude_linear)
    output_df_to_excel(df, cal_pred, result_df)

# Generate profiles
def generate_profiles():
    """
    Generate profiles for patients.
    
    Details:
    Define parameters and lists. Loop through patients to clean data, select data for calibration and efficacy-driven dosing, keep patients with
    sufficient dose-response pairs and predictions for each method, then apply method.
    
    Outputs:
    - patients_to_exclude_linear: list of patients to exclude for linear methods.
    - list_of_patient_df: list of dataframes of cleaned patient data.
    - list_of_cal_pred_df: list of dataframes of data for calibration and efficacy-driven dosing, that has sufficient dose-response pairs and predictions
    - list_of_result_df: list of dataframes with results after applying methods.
    """

    # Profile Generation
    input_file = 'Retrospective Liver Transplant Data.xlsx'
    rows_to_skip = 17

    # Get list of patients/sheet names
    list_of_patients = get_sheet_names(input_file)

    # Define lists
    list_of_patient_df = []
    list_of_cal_pred_df = []
    list_of_result_df = []
    patients_to_exclude_linear = []

    number_of_patients = 0

    for patient in list_of_patients:

        # Create and clean patient dataframe        
        df = pd.read_excel(input_file, sheet_name=patient, skiprows=rows_to_skip)
        df = clean_data(df)
        df = keep_ideal_data(df, patient, list_of_patient_df)
        
        # Counter for number of patients
        number_of_patients = number_of_patients + 1

        # Select data for calibration and efficacy-driven dosing
        cal_pred_linear, patients_to_exclude_linear = cal_pred_data(df, patient, patients_to_exclude_linear)

        # Keep patients with sufficient dose-response pairs and predictions for each method
        cal_pred, list_of_cal_pred_df = keep_target_patients(patient, patients_to_exclude_linear, 
                                                         cal_pred_linear, list_of_cal_pred_df)   

        # Apply methods
        list_of_result_df = apply_methods(cal_pred, patient, patients_to_exclude_linear,
                  cal_pred_linear, list_of_result_df)

    return patients_to_exclude_linear, list_of_patient_df, list_of_cal_pred_df, list_of_result_df

def print_patients_to_exclude(patients_to_exclude_linear):
    """Print patients to exclude"""        
    patients_to_exclude_linear = sorted(set(patients_to_exclude_linear))
    print(f"Patients to exclude: {patients_to_exclude_linear}")

def join_dataframes(list_of_patient_df, list_of_cal_pred_df, list_of_result_df):
    """Join dataframes from individual patients"""
    df = pd.concat(list_of_patient_df)
    df.patient = df.patient.apply(int)
    df.reset_index(inplace=True, drop=True)
    cal_pred = pd.concat(list_of_cal_pred_df)
    cal_pred.patient = cal_pred.patient.apply(int)
    result_df = pd.concat(list_of_result_df)
    result_df = format_result_df(cal_pred, result_df)
    
    return df, cal_pred, result_df

def output_df_to_excel(df, cal_pred, result_df):
    """Output dataframes to excel as individual sheets"""
    file_name = 'CURATE_results.xlsx'

    with pd.ExcelWriter(file_name) as writer:
        df.to_excel(writer, sheet_name='clean', index=False)
        cal_pred.to_excel(writer, sheet_name='calibration_and_efficacy_driven', index=False)
        result_df.to_excel(writer, sheet_name='result', index=False)

# Create patient dataframe        
def get_sheet_names(input_file):
    """ Get sheet names which are also patient names """
    wb = load_workbook(input_file, read_only=True)
    patient_list = wb.sheetnames
    wb.close()
    return patient_list

def clean_data(df):
    """ 
    Keep target columns from excel, shift tac level one cell up, remove "mg"/"ng" 
    from dose, replace NaN with 0 in dose.
    
    Input:
    df - dataframe from each sheet
    patient - patient number from list of patients
    
    Output:
    df - cleaned dataframe        
    """
    dose_string = "2nd Tac dose (pm)"

    # Keep target columns
    df = df[["Day #", "Tac level (prior to am dose)", dose_string]]

    # Shift tac level one cell up to match dose-response to one day
    df[dose_string] = df[dose_string].shift(1)

    # Remove "mg"/"ng" from dose
    df[dose_string] = df[dose_string].astype(str).str.replace('mgq', '')
    df[dose_string] = df[dose_string].astype(str).str.replace('mg', '')
    df[dose_string] = df[dose_string].astype(str).str.replace('ng', '')
    df[dose_string] = df[dose_string].astype(str).str.strip()
    df[dose_string] = df[dose_string].astype(float)

    first_day_of_dosing = df['Day #'].loc[~df[dose_string].isnull()].iloc[0]

    # Keep data from first day of dosing
    df = df[df['Day #'] >= first_day_of_dosing].reset_index(drop=True)

    # Set the first day of dosing as day 2 (because first dose received on day 1)
    for i in range(len(df)):
        df.loc[i, 'Day #'] = i + 2
    
    return df

def keep_ideal_data(df, patient, list_of_patient_df):
    """
    Remove rows with non-ideal data, including missing tac level and/or tac dose, 
    <2 tac level, multiple blood draws. Then keep the longest consecutive chunk
    based on number of days. 

    Input: Dataframe of individual patient
    Output: Dataframe with the longest chunk of consecutive ideal data.
    """

    dose_string = "2nd Tac dose (pm)"

    # Create boolean column of data to remove
    # Including NA, <2 tac level, multiple blood draws
    df['non_ideal'] = (df.isnull().values.any(axis=1))  | \
                               (df['Tac level (prior to am dose)'] == '<2') | \
                               (df["Tac level (prior to am dose)"].astype(str).str.contains("/"))

    # # Set boolean for non_ideal as True if all dose including and above current row is 0
    # for i in range(len(df)):
    #     if (df.loc[0:i, dose_string] == 0).all():
    #         df.loc[i, 'non_ideal'] = True

    # Create index column
    df.reset_index(inplace=True) 

    try:
        # Cum sum
        df['cum_sum'] = df.non_ideal.cumsum()
        df_agg = df[df.non_ideal==False].groupby('cum_sum').agg({'index': ['count', 'min', 'max']})
        df_agg = df_agg[df_agg['index']['count']==df_agg['index']['count'].max()]

        # Find index of data to keep
        if len(df_agg) > 1:
            print('there are multiple sets of consecutively ideal data')
            df_agg = df_agg.loc[0,:]
        min_idx = df_agg['index','min'].squeeze()
        max_idx = df_agg['index','max'].squeeze()

        # Keep longest consecutive set of ideal data
        df = df.iloc[min_idx: max_idx+1, :]
    except:
        print(f"no ideal data for patient {patient}")

    # Format patient dataframe
    df['patient'] = patient
    df = df[['Day #', 'Tac level (prior to am dose)', dose_string, 'patient']]
    df.columns = ['day', 'response', 'dose', 'patient']
    list_of_patient_df.append(df)

    return df

def cal_pred_data(df, patient, patients_to_exclude):
    """
    Find calibration points and combine calibration data with the rest of data,
    to perform CURATE methods later.

    Input: 
    df - dataframe
    patient - string of patient number

    Output: 
    Dataframe of calibration data and data for subsequent predictions.
    Print patients with insufficient data for calibration and <3 predictions

        """
    # Create index column
    df = df.reset_index(drop=False) 

    # Create boolean column to check if tac dose diff from next row
    df['diff_from_next'] = \
            (df['dose'] != df['dose'].shift(-1))

    # Find indexes of last rows of first 2 unique doses
    last_unique_doses_idx = [i for i, x in enumerate(df.diff_from_next) if x]

    # Create boolean column to check if tac dose diff from previous row
    df['diff_from_prev'] = \
            (df['dose'] != df['dose'].shift(1))

    # Find indexes of first rows of third unique dose
    first_unique_doses_idx = [i for i, x in enumerate(df.diff_from_prev) if x]

    # The boolean checks created useless index, diff_from_next and diff_from_prev columns,
    # drop them
    df = df.drop(['index', 'diff_from_next', 'diff_from_prev'], axis=1)

    # Combine calibration and prediction rows
    cal_pred = pd.DataFrame()

    if df['dose'].nunique() > 1:
        first_cal_point = pd.DataFrame(df.iloc[last_unique_doses_idx[0],:]).T
        second_cal_point = pd.DataFrame(df.iloc[first_unique_doses_idx[1],:]).T
        rest_of_data = df.iloc[first_unique_doses_idx[1]+1:,:]
        cal_pred = pd.concat([first_cal_point, second_cal_point, 
                                    rest_of_data]).reset_index(drop=True)
    else:
        patients_to_exclude.append(str(patient))
        print(f"Patient #{patient} has insufficient unique dose-response pairs for calibration (for linear)!")

    # Print error msg if number of predictions is less than 3
    if (len(cal_pred) - (2) < 3) and (df['dose'].nunique() > 1):
        patients_to_exclude.append(str(patient))
        print(f"Patient #{patient} has insufficient/<3 predictions ({len(cal_pred) - (2)} predictions)!")

    # Add "type" column
    cal_pred['type'] = 'linear'

    return cal_pred, patients_to_exclude

def keep_target_patients(patient, patients_to_exclude_linear, cal_pred_linear, list_of_cal_pred_df):
    """
    Keep patient data with sufficient dose-response pairs and predictions
    
    Details:
    There are 4 scenarios depending on inclusion/exlusion of patient for linear methods.
    Append patient data to list_of_cal_pred_df only if there's sufficient dose-response pairs and predictions.
    
    Input:
    patient
    patients_to_exclude_linear - list of patients to be excluded from linear methods
    cal_pred_linear - individual patient dataframe with calibration and efficacy-driven dosing data for linear methods
    list_of_cal_pred_df - list of all cal_pred dataframes of each patient from first to current patient in loop
    
    Output:
    cal_pred
        - individual patient dataframe of calibration and efficacy-driven dosing, 
          that has sufficient dose-response pairs and predictions
    list_of_cal_pred_df 
        - list of all cal_pred dataframes of each patient 
          from first to current patient in loop
    """
    cal_pred = pd.DataFrame()
    if (patient not in patients_to_exclude_linear):
        cal_pred = cal_pred_linear
        list_of_cal_pred_df.append(cal_pred)
    elif patient in patients_to_exclude_linear:
        list_of_cal_pred_df.append(cal_pred)
    elif patient not in patients_to_exclude_linear:
        cal_pred = cal_pred_linear
        list_of_cal_pred_df.append(cal_pred)
    else: cal_pred = pd.DataFrame()

    cal_pred.reset_index(inplace=True, drop=True)

    return cal_pred, list_of_cal_pred_df

# Prepare patient dataframe for prediction and apply method
def apply_methods(cal_pred, patient, patients_to_exclude_linear,
                  cal_pred_linear, list_of_result_df):
    
    """
    If cal_pred is filled, create result dataframe and apply all methods.
    
    Input:
    cal_pred - dataframe of cal_pred_linear that satisfy minimum criteria.
    patient
    patients_to_exclude_linear - list of patients to exclude for linear methods
    cal_pred_linear - individual patient dataframe with calibration and efficacy-driven dosing data for linear methods
    list_of_result_df - list of result dataframe from each patient
    
    Output:
    list_of_result_df - list of result dataframe from each patient
    """
    
    if len(cal_pred) != 0:

        # Create result dataFrame
        max_count_input = len(cal_pred_linear)

        col_names = ['patient', 'method', 'pred_day'] + \
                    ['fit_dose_' + str(i) for i in range(1, max_count_input + 1)] + \
                    ['fit_response_' + str(i) for i in range(1, max_count_input + 1)] + \
                    ['day_' + str(i) for i in range(1, max_count_input + 1)] + \
                    ['weight_' + str(i) for i in range(1, max_count_input + 1)] + \
                    ['half_life', 'dose', 'response', 'prev_coeff_2x', 'prev_coeff_1x', 'prev_coeff_0x',\
                     'prev_deviation', 'coeff_2x', 'coeff_1x', 'coeff_0x', 'prediction', 'deviation',
                     'abs_deviation']
        result = pd.DataFrame(columns=col_names)    

        if patient not in patients_to_exclude_linear:
            deg = 1

            list_of_result_df = RW(deg, cal_pred_linear, result, 'L_RW_wo_origin', list_of_result_df)

    return list_of_result_df

def RW(deg, cal_pred, result, method_string, list_of_result_df):
    """
    Execute rolling window approach for each variation of origin inclusion and type. 

    Details:
    Add a fresh line to result dataframe for each prediction and fill in the line with patient, method,
    prediction day. Fill in the dose-response pairs and days of the pairs that are 
    required to fit the model, based on last deg + 1 unique dose-response pairs, according to the origin 
    inclusion chosen. Fill in the dose-response pair of the day of the prediction to calculate the 
    deviation of the prediction. Fit the model. Calculate the prediction, deviation, and absolute deviation, 
    and fill in the results to the result dataframe.
    
    Input:
    deg - degree of polynomial fit, 1 for linear, 2 for quadratic
    cal_pred - dataframe of calibration and efficacy-driven dosing data, 
                cal_pred_linear for linear, cal_pred_quad for quad
    result - dataframe of results
    method_string - string of method
    list_of_result_df - list of result dataframe from each patient

    Output:
    list_of_result_df
    """
    values = []
    indices = []
    
    j = 0

    result = result[0:0]

    # Loop through rows
    for i in range(0, len(cal_pred)-1):

        # Define new dose
        new_dose = cal_pred.loc[i, 'dose']

        # Add values and indices of first three unique doses into list
        if len(values) < deg + 1:

            # Check if new dose in values list
            # If not, append new dose into value list and append index into indices list
            if new_dose not in values:
                values.append(new_dose)
                indices.append(i)

        else:
            # Check if new dose in value list
            # If yes: remove value and index of repeat, and append new dose and index to list.
            if new_dose in values:
                repeated_dose_index = values.index(new_dose)
                values.pop(repeated_dose_index)
                indices.pop(repeated_dose_index)

                values.append(new_dose)
                indices.append(i)

            # If no: remove the first element of list and append value and index of new dose
            else:
                values.pop(0)
                indices.pop(0)

                values.append(new_dose)
                indices.append(i)

        # Prepare dataframe of dose-response pairs for prediction
        counter_for_origin_int = 1
        if len(indices) == deg + 1:            

            result.loc[j, 'patient'] = cal_pred.loc[i + 1, 'patient']
            result.loc[j, 'method'] = method_string
            result.loc[j, 'pred_day'] = cal_pred.loc[i + 1, 'day']
            
            # Fill in dose, response, day to fit model
            cal_pred_string = ['dose', 'response', 'day']
            result_string = ['fit_dose', 'fit_response', 'day']
            for string_num in range(0,len(cal_pred_string)):
                result.loc[j, result_string[string_num] + '_1'] = cal_pred.loc[indices[0], cal_pred_string[string_num]]
                result.loc[j, result_string[string_num] + '_2'] = cal_pred.loc[indices[1], cal_pred_string[string_num]]
                if deg == 2:
                    result.loc[j, result_string[string_num] + '_3'] = cal_pred.loc[indices[2], cal_pred_string[string_num]]

            # Fill in new dose and response of prediction day
            result.loc[j, 'dose'] = cal_pred.loc[i + 1, 'dose']
            result.loc[j, 'response'] = cal_pred.loc[i + 1, 'response']
            
            # Curve fit equation
            X = result.loc[j, 'fit_dose_1':'fit_dose_' + str(deg+1)].to_numpy()
            y = result.loc[j, 'fit_response_1':'fit_response_' + str(deg+1)].to_numpy()

            poly_reg = PolynomialFeatures(degree=deg)
            X = X.reshape(-1,1)
            X = poly_reg.fit_transform(X)
            fitted_model = LinearRegression(fit_intercept=False).fit(X, y)
            result.loc[j, 'coeff_0x'] = fitted_model.coef_[0]
            result.loc[j, 'coeff_1x'] = fitted_model.coef_[1]
                
            # Calculate prediction and deviation
            prediction = fitted_model.predict(poly_reg.fit_transform([[cal_pred.loc[i, 'dose']]]))[0]

            result.loc[j, 'prediction'] = prediction
            deviation = cal_pred.loc[i + 1, 'response'] - prediction
            result.loc[j, 'deviation'] = deviation
            abs_deviation = abs(deviation)
            result.loc[j, 'abs_deviation'] = abs_deviation

            counter_for_origin_int = counter_for_origin_int + 1
            
        j = j + 1
    
    list_of_result_df.append(result)

    return list_of_result_df

# Combine dataframes of individual patients
def format_result_df(cal_pred, result_df):
    """
    Rerrange column names, make patient and prediction day series of integers, sort, reset index
    
    Input: 
    cal_pred - dataframe of calibration and efficacy-driven dosing data for all patients
    result_df - results after applying all methods to patients
    
    Output:
    result_df - Formatted result_df
    """
    max_count_input = cal_pred[cal_pred['type']=='linear'].groupby('patient').count().max()['dose']
    col_names = ['patient', 'method', 'pred_day'] + \
                ['fit_dose_' + str(i) for i in range(1, max_count_input + 1)] + \
                ['fit_response_' + str(i) for i in range(1, max_count_input + 1)] + \
                ['day_' + str(i) for i in range(1, max_count_input + 1)] + \
                ['weight_' + str(i) for i in range(1, max_count_input + 1)] + \
                ['half_life', 'dose', 'response', 'prev_coeff_2x', 'prev_coeff_1x', 'prev_coeff_0x',\
                 'prev_deviation', 'coeff_2x', 'coeff_1x', 'coeff_0x', 'prediction', 'deviation',
                 'abs_deviation']
    result_df = result_df[col_names]
    result_df.patient = result_df.patient.apply(int)
    result_df.pred_day = result_df.pred_day.apply(int)
    result_df.sort_values(['patient', 'method', 'half_life', 'pred_day'], inplace=True)
    result_df = result_df.reset_index(drop=True)
    
    return result_df

# Consolidate all patient data and label those assessed as ideal or non-ideal for analysis
def all_data():
    """
    Clean raw data and label which are ideal or non-ideal.
    
    Output: 
    - Dataframe of all cleaned raw data with label of ideal/non-ideal.
    - 'all_data.xlsx' with dataframe
    """
    dose_string = "Eff 24h Tac Dose"
    result_file = "CURATE_results.xlsx"

    # Create dataframe from all sheets
    list_of_patients = find_list_of_patients()

    df = pd.DataFrame()

    for patient in list_of_patients:
        patient_df = pd.read_excel('Retrospective Liver Transplant Data.xlsx', sheet_name=patient, skiprows=17)
        patient_df['patient'] = patient

        # Subset dataframe
        patient_df = patient_df[['Day #', 'Tac level (prior to am dose)', dose_string, 'patient']]

        # Shift dose column
        patient_df[dose_string] = patient_df[dose_string].shift(1)

        # Remove "mg"/"ng" from dose
        patient_df[dose_string] = patient_df[dose_string].astype(str).str.replace('mgq', '')
        patient_df[dose_string] = patient_df[dose_string].astype(str).str.replace('mg', '')
        patient_df[dose_string] = patient_df[dose_string].astype(str).str.replace('ng', '')
        patient_df[dose_string] = patient_df[dose_string].astype(str).str.strip()
        patient_df[dose_string] = patient_df[dose_string].astype(float)

        first_day_of_dosing = patient_df['Day #'].loc[~patient_df[dose_string].isnull()].iloc[0]

        # Keep data from first day of dosing
        patient_df = patient_df[patient_df['Day #'] >= first_day_of_dosing].reset_index(drop=True)

        # Set the first day of dosing as day 2 (because first dose received on day 1)
        for i in range(len(patient_df)):
            patient_df.loc[i, 'Day #'] = i + 2

        df = pd.concat([df, patient_df])

    df.reset_index(drop=True)
    df['patient'] = df['patient'].astype(int)

    # Rename columns
    df = df.rename(columns={'Day #':'day', 'Tac level (prior to am dose)':'response', dose_string:'dose'})

    # Import output dataframe from 'clean'
    ideal_df = pd.read_excel(result_file, sheet_name='clean')

    # Add ideal column == TRUE
    ideal_df['ideal'] = True

    # Subset columns
    ideal_df = ideal_df[['day', 'patient', 'ideal']]

    # Merge dataframes
    combined_df = df.merge(ideal_df, how='left', on=['day', 'patient'])

    # Fill in ideal column with False if NaN
    combined_df['ideal'] = combined_df['ideal'].fillna(False)

    for i in range(len(combined_df)):
        # Find index of patient in list_of_patients
        index = list_of_patients.index(str(combined_df.patient[i]))

    # Add column 'dose'
    combined_df['dose'] = combined_df['dose'].astype(float)

    # Clean up response column
    for i in range(len(combined_df)):

        # For response column, if there is a '/', take first value
        if '/' in str(combined_df.response[i]):
            combined_df.loc[i, 'response'] = combined_df.response[i].split('/')[0]
        else: 
            combined_df.loc[i, 'response'] = combined_df.response[i]

        # If response is <2, which means contain '<', label as NaN
        if '<' in str(combined_df.response[i]):
            combined_df.loc[i, 'response'] = np.nan

    # Export to excel
    combined_df.to_excel(r'all_data.xlsx', index = False, header=True)
    
    return combined_df

# Calculate dose recommendations
def dose_recommendation_results():
    """
    Compute dose recommendations after fitting the training data with L_RW_wo_origin.

    Output: dose_recommendations.xlsx containing dose recommendation results
    """
    minimum_capsule = 0.5

    result_file = "CURATE_results.xlsx"
    
    df = pd.read_excel(result_file, sheet_name='result')
    df = df[df.method=='L_RW_wo_origin'].reset_index(drop=True)

    # Interpolate
    for i in range(len(df)):
        # Create function
        coeff = df.loc[i, 'coeff_2x':'coeff_0x'].apply(float).to_numpy()
        coeff = coeff[~np.isnan(coeff)]
        p = np.poly1d(coeff)
        x = np.linspace(0, 8)
        y = p(x)

        # Check for duplicates, which will occur if coeff_1x is very close to 0, and
        # will cause RuntimeError for interp1d. Hence, set interpolated doses to the intercept,
        # also known as coeff_0x
        dupes = [x for n, x in enumerate(y) if x in y[:n]]
        if len(dupes) != 0:
            df.loc[i, 'interpolated_dose_8'] = df.loc[i, 'coeff_0x']
            df.loc[i, 'interpolated_dose_9'] = df.loc[i, 'coeff_0x']
            df.loc[i, 'interpolated_dose_10'] = df.loc[i, 'coeff_0x']

        else:
            f = interpolate.interp1d(y, x, fill_value='extrapolate')

            df.loc[i, 'interpolated_dose_8'] = f(8)
            df.loc[i, 'interpolated_dose_9'] = f(9)
            df.loc[i, 'interpolated_dose_10'] = f(10)

    # Possible dose
    df['possible_doses'] = ""
    for i in range(len(df)):
        # Find minimum dose recommendation by mg
        min_dose_mg = math.ceil(min(df.interpolated_dose_8[i], df.interpolated_dose_10[i]) * 2) / 2

        # Find maximum dose recommendation by mg
        max_dose_mg = math.floor(max(df.interpolated_dose_8[i], df.interpolated_dose_10[i]) * 2) / 2

        # Between and inclusive of min_dose_mg and max_dose_mg,
        # find doses that are multiples of 0.5 mg
        possible_doses = np.arange(min_dose_mg, max_dose_mg + minimum_capsule, minimum_capsule)
        possible_doses = possible_doses[possible_doses % minimum_capsule == 0]

        if possible_doses.size == 0:
            possible_doses = np.array(min(min_dose_mg, max_dose_mg))

        # Add to column of possible doses
        df.at[i, 'possible_doses'] = possible_doses

        # Add to column of dose recommendation with lowest out of possible doses
        df.loc[i, 'dose_recommendation'] = possible_doses if (possible_doses.size == 1) else min(possible_doses)

    # Resulting response with dose recommendation
    df['predicted_response_after_recommended_dose'] = df['coeff_1x'] * df['dose_recommendation'] + df['coeff_0x']

    # Export to excel
    df.to_excel(r'dose_recommendations.xlsx', index = False, header=True)

    return df

if __name__ == '__main__':
    # Get arguments
    import argparse
    parser = argparse.ArgumentParser(description='Implement CURATE models and retrieve results')
    args = parser.parse_args()
    
    # Implement CURATE.AI models
    print('implementing CURATE.AI models...')
    original_stdout = sys.stdout
    with open('patients_to_exclude.txt', 'w') as f:
        sys.stdout = f
        execute_CURATE()
    sys.stdout = original_stdout
        
    # Consolidate all patient data, label them as ideal or non-ideal for analysis
    print('consolidating all patient data and labeling them as ideal or non-ideal for analysis...')   
    all_data()
    
    # Dose recommendations
    print('computing CURATE.AI dose recommendations...')
    dose_recommendation_results()
    print('end of code')