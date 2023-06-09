import seaborn as sns
import pandas as pd
import numpy as np
from sklearn.metrics import mean_squared_error
import math
import matplotlib.pyplot as plt
from matplotlib import colors
from matplotlib.pyplot import cm
from matplotlib.patches import Patch
import matplotlib as mpl
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook
import sys
from scipy import stats
from scipy.stats import mannwhitneyu
from scipy.stats import levene
from scipy.stats import wilcoxon
from scipy.stats import bartlett
from statistics import *
from scipy import interpolate

from implement_CURATE import *

# Define file names
result_file = 'CURATE_results.xlsx'
raw_data_file = 'data_retro_ped_liver_transplant.xlsx'
all_data_file = 'all_data.xlsx'

# Define clinically relevant parameters
low_dose_upper_limit = 2
medium_dose_upper_limit = 4
overprediction_limit = -1.5
underprediction_limit = 2
max_dose_recommendation = 8
min_dose_recommendation = 0
minimum_capsule = 0.5
therapeutic_range_upper_limit = 10
therapeutic_range_lower_limit = 8
dosing_strategy_cutoff = 0.4
acceptable_tac_upper_limit = 12
acceptable_tac_lower_limit = 6.5

# Patient population
def percentage_of_pts_that_reached_TR_per_dose_range(all_data_file=all_data_file):
    """Find percentage of patients that reached TR at the same dose range."""
    # Filter doses that reached TR
    df = pd.read_excel(all_data_file)
    df = df[df['dose'].notna()].reset_index(drop=True)
    for i in range(len(df)):
        if (df.response[i] >= therapeutic_range_lower_limit) & (df.response[i] <= therapeutic_range_upper_limit):
            df.loc[i, 'TR'] = True
        else:
            df.loc[i, 'TR'] = False

    df = df[df.TR==True].reset_index(drop=True)

    # Dose range grouped by patients
    for i in range(len(df)):
        if df.dose[i] < low_dose_upper_limit:
            df.loc[i, 'dose_range'] = 'low'
        elif df.dose[i] < medium_dose_upper_limit:
            df.loc[i, 'dose_range'] = ' medium'
        else:
            df.loc[i, 'dose_range'] = 'high'

    # Dose range that reached TR
    df = df.groupby('patient')['dose_range'].unique().astype(str).reset_index(name='dose_range_in_TR')

    # Percentage of patients by TR dose range
    df = df.groupby('dose_range_in_TR')['dose_range_in_TR'].apply(lambda x: x.count()/len(df.patient.unique())*100).reset_index(name='perc_pts')

    # Format percentage
    df['perc_pts'] = df['perc_pts'].apply(lambda x: "{:.2f}".format(x))

    return df
    
def patient_population_values():
    """
    Print out results of the following into a text file 'patient_population_values.txt'
    1. Response
    2. % of days within therapeutic range
    3. % of participants that reached therapeutic range within first week
    4. Day where patient first achieved therapeutic range
    5. Dose administered by mg
    """
    original_stdout = sys.stdout
    with open('patient_population_values.txt', 'w') as f:
        sys.stdout = f
        
        data = fig_2(plot=False)

        # 1. Response
        result_and_distribution(data.response, '1. Response')

        # 2. % of days within therapeutic range

        # Drop rows where response is NaN
        data = data[data.response.notna()].reset_index(drop=True)

        # Add therapeutic range column
        for i in range(len(data)):
            if (data.response[i] >= therapeutic_range_lower_limit) & (data.response[i] <= therapeutic_range_upper_limit):
                data.loc[i, 'therapeutic_range'] = True
            else:
                data.loc[i, 'therapeutic_range'] = False

        perc_therapeutic_range = data.groupby('patient')['therapeutic_range'].apply(lambda x: x.sum()/x.count()*100)
        perc_therapeutic_range = perc_therapeutic_range.to_frame().reset_index()

        # Result and distribution
        result_and_distribution(perc_therapeutic_range.therapeutic_range, '2. % of days within therapeutic range')

        # 3. % of participants that reached therapeutic range within first week
        first_week_df = data.copy()
        first_week_df = first_week_df[first_week_df['Tacrolimus trough levels (TTL)']=='Within the therapeutic range'].reset_index(drop=True)
        first_week_df = (first_week_df.groupby('patient')['Day'].first() <= 7).to_frame().reset_index()
        result = first_week_df.Day.sum()/first_week_df.Day.count()*100

        print(f'3. % of participants that reached therapeutic range within first week:\n{result:.2f}%,\
        {first_week_df.Day.sum()} out of 16 patients\n')

        # 4. Day where patient first achieved therapeutic range
        first_TR_df = data.copy()
        first_TR_df = first_TR_df[first_TR_df['Tacrolimus trough levels (TTL)']=='Within the therapeutic range'].reset_index(drop=True)
        first_TR_df = first_TR_df.groupby('patient')['Day'].first().to_frame().reset_index()

        # Result and distribution
        result_and_distribution(first_TR_df.Day, '4. Day where patient first achieved therapeutic range')

        # 5. Dose administered by mg
        dose_df = data.copy()
        result_and_distribution(dose_df.dose, '5. Dose administered')

    sys.stdout = original_stdout

# Fig 2
def fig_2(file_string=all_data_file, plot=False):
    """Scatter plot of inidividual profiles, longitudinally, and response vs dose"""
    
    file_string = all_data_file

    # Plot individual profiles
    dat = pd.read_excel(file_string)

    # Create within-range column for color
    dat['within_range'] = (dat.response <= therapeutic_range_upper_limit) & (dat.response >= therapeutic_range_lower_limit)

    # Create low/med/high dose column
    dat['dose_range'] = ""
    for i in range(len(dat)):
        if np.isnan(dat.dose[i]):
             dat.loc[i, 'dose_range'] = 'Unavailable'
        elif dat.dose[i] < low_dose_upper_limit:
            dat.loc[i, 'dose_range'] = 'Low'
        elif dat.dose[i] < medium_dose_upper_limit:
            dat.loc[i, 'dose_range'] = 'Medium'
        else:
            dat.loc[i, 'dose_range'] = 'High'

    # Rename columns and entries
    new_dat = dat.copy()
    new_dat = new_dat.rename(columns={'within_range':'Tacrolimus trough levels (TTL)'})
    new_dat['Tacrolimus trough levels (TTL)'] = new_dat['Tacrolimus trough levels (TTL)'].map({True:'Within the therapeutic range', False: 'Outside of the therapeutic range'})
    new_dat = new_dat.rename(columns={'dose_range':'Dose range', 'day':'Day'})

    print(new_dat.columns)

    if plot == True:

        # Add fake row with empty data under response to structure legend columns
        new_dat.loc[len(new_dat.index)] = [2, 5, 0.5, 1, True, "", "Low"]
        new_dat.loc[len(new_dat.index)] = [2, 5, 0.5, 1, True, "", "Low"]
        
        # Plot tac levels by day
        sns.set(font_scale=1.5, rc={"figure.figsize": (16,10), "xtick.bottom" : True, "ytick.left" : True}, style='white')

        g = sns.relplot(data=new_dat, x='Day', y='response', hue='Tacrolimus trough levels (TTL)', col='patient', col_wrap=4, style='Dose range',
                height=3, aspect=1,s=100, palette=['tab:blue','tab:orange','white'], 
                style_order=['Low', 'Medium', 'High', 'Unavailable'], zorder=2, edgecolor=None)

        # g = sns.relplot(data=new_dat[new_dat['Dose range']=='Low'], x='Day', y='response', hue='TTL', col='patient', col_wrap=4, style='o',
        # height=3, aspect=1,s=100, palette=['tab:blue','tab:orange','white','white'])
        
        # Add gray region for therapeutic range
        for ax in g.axes:
            ax.axhspan(therapeutic_range_lower_limit, therapeutic_range_upper_limit, facecolor='grey', alpha=0.2, zorder=1)
        
        g.set_titles('Patient {col_name}')
        g.set_ylabels('TTL (ng/ml)')
        g.set(yticks=np.arange(0,math.ceil(max(new_dat.response)),4),
            xticks=np.arange(0, max(new_dat.Day+2), step=5))
        
        # Move legend below plot
        sns.move_legend(g, 'center', bbox_to_anchor=(0.245,-0.08), ncol=2)
        legend_elements = [Patch(facecolor='grey', edgecolor='grey',
                            label='Region within therapeutic range', alpha=.2)]
        legend2 = plt.legend(handles=legend_elements, bbox_to_anchor=(-1.0,-0.4), loc='upper left', frameon=False)

        # plt.savefig('fig_2_' + dose + '.png', dpi=1000, facecolor='w', bbox_inches='tight')
        plt.savefig('fig_2.svg', format='svg', dpi=1000, facecolor='w', bbox_inches='tight')

        # Remove fake row before end of function
        new_dat = new_dat[:-1]

    return new_dat

# Technical performance metrics
def technical_performance_metrics():
    """
    Print the following technical performance metrics
    1. Prediction erorr
    2. Absolute prediction error
    3. RMSE
    """
    original_stdout = sys.stdout
    with open('technical_perf_metrics.txt', 'w') as f:
        sys.stdout = f

        df = pd.read_excel(result_file, sheet_name='result')

        # Subset method
        df = df[df.method=='L_RW_wo_origin'].reset_index(drop=True)

        # 1. Prediction error
        result_and_distribution(df.deviation, '1. Prediction error')

        print('*for comparison with non-parametric abs prediction error,')
        median_IQR_range(df.deviation)

        # 2. Absolute prediction error
        result_and_distribution(df.abs_deviation, '2. Absolute prediction error')

        # 3. RMSE
        RMSE = (math.sqrt(mean_squared_error(df.response, df.prediction)))
        print(f'3. RMSE: {RMSE:.2f}\n')

        ## Compare medians between training and test sets

    sys.stdout = original_stdout

# Clinically relevant performance metrics
def clinically_relevant_performance_metrics():
    """Clinically relevant performance metrics. 
    Find the percentage of predictions within clinically acceptable prediction error,
    percentage of overpredictions, and percentage of underpredictions. Print the results
    into a text file.
    """
    # Uncomment to write output to txt file
    # file_path = 'Clinically relevant performance metrics.txt'
    # sys.stdout = open(file_path, "w")
    
    original_stdout = sys.stdout
    with open('clinically_relevant_perf_metrics.txt', 'w') as f:
        sys.stdout = f

        df = pd.read_excel(result_file, sheet_name='result')
        df = df[df.method=='L_RW_wo_origin'].reset_index()
        df = df[['patient', 'pred_day', 'deviation']]   

        # Percentage of predictions within clinically acceptable prediction error
        acceptable = df[(df.deviation > overprediction_limit) & (df.deviation < underprediction_limit)].reset_index(drop=True)
        perc_of_clinically_acceptable_pred = len(acceptable)/len(df)*100

        # Percentage of overpredictions
        overpred = df[df.deviation < overprediction_limit]
        perc_of_overpred = len(overpred)/len(df)*100

        # Percentage of underpredictions
        underpred = df[df.deviation > underprediction_limit]
        perc_of_underpred = len(underpred)/len(df)*100

        print(f'% of predictions within clinically acceptable prediction error: {perc_of_clinically_acceptable_pred:.2f}%, n = {len(df)}\n')
        print(f'% of predictions which are overpredictions: {perc_of_overpred:.2f}%, n = {len(df)}\n')
        print(f'% of predictions which are underpredictions: {perc_of_underpred:.2f}%, n = {len(df)}')

    sys.stdout = original_stdout

# Fig 4
def values_in_clinically_relevant_flow_chart():
    """
    Calculate values for clinically relevant flow chart, in the flow chart boxes, and in additional information

    Output: 
    - Printed values for flow chart boxes and additional information
    - Final dataframe with remaining predictions after all exclusions
    """
    original_stdout = sys.stdout

    file_string = 'clinically_relevant_flow_chart.txt'
    with open(file_string, 'w') as f:
        sys.stdout = f

        result_file = 'CURATE_results.xlsx'

        df = create_df_for_CURATE_assessment()

        # 1. Calculate values for clinially relevant flow chart (step 1 of 2)

        total_predictions = len(df)
        unreliable = len(df[df.reliable==False])

        # Keep reliable predictions
        df = df[df.reliable==True].reset_index(drop=True)

        reliable = len(df)
        inaccurate = len(df[df.accurate==False])

        # Keep accurate predictions
        df = df[df.accurate==True].reset_index(drop=True)

        reliable_accurate = len(df)

        print(f'Flowchart numbers:\ntotal_predictions {total_predictions} | unreliable {unreliable} | reliable {reliable} | inaccurate {inaccurate} | reliable_accurate {reliable_accurate}')

        # 2. Calculate values for additional information

        reliable_accurate_actionable = len(df[df.actionable==True])

        # Keep actionable predictions
        df = df[df.actionable==True].reset_index(drop=True)

        reliable_accurate_actionable_diff_dose = len(df[df.diff_dose==True])

        # Keep diff dose predictions
        df = df[df.diff_dose==True].reset_index(drop=True)

        reliable_accurate_actionable_diff_dose_non_therapeutic_range = len(df[df.therapeutic_range==False])

        # Keep non therapeutic range
        df = df[df.therapeutic_range==False].reset_index(drop=True)

        print(f'\nAdditional information:\nreliable_accurate_actionable {reliable_accurate_actionable} out of {reliable_accurate} |\n\
        reliable_accurate_actionable_diff_dose {reliable_accurate_actionable_diff_dose} out of {reliable_accurate_actionable} |\n\
        reliable_accurate_actionable_diff_dose_non_therapeutic_range {reliable_accurate_actionable_diff_dose_non_therapeutic_range} out of {reliable_accurate_actionable_diff_dose}')

        try:
            # Add column for difference in doses recommended and administered
            df['diff_dose_recommended_and_administered'] = df['dose_recommendation'] - df['dose']
            df['diff_dose_recommended_and_administered'] = df['diff_dose_recommended_and_administered'].astype(float)
            result_and_distribution(df.diff_dose_recommended_and_administered, 'Dose recommended minus administered')
        except:
            print('dose difference calculation failed')
        

    sys.stdout = original_stdout

    return df

# Fig 5
def fig_5a(plot=True):
    case_series_patient_num = 5
    case_series_patient_journey(case_series_patient_num=case_series_patient_num, all_data_file=all_data_file)

def case_series_patient_journey(case_series_patient_num, all_data_file, plot=True):
    # SOC data
    case_series_patient = pd.read_excel(all_data_file)
    case_series_patient = case_series_patient[case_series_patient.patient==case_series_patient_num]
    case_series_patient = case_series_patient[['day', 'response']].reset_index(drop=True)
    
    if case_series_patient_num==4:
        case_series_patient = case_series_patient.iloc[:13,:] # Added this line for patient 4 to stop from the row where the rest of the days have NaN response

    # CURATE data
    df = create_df_for_CURATE_assessment()
    df = df[df.patient==case_series_patient_num].reset_index(drop=True)

    # Merge both dataframes
    df = df.rename(columns={'pred_day':'day'})
    combined_df = case_series_patient.merge(df[['day', 'projected_response']], how='left', on='day')
    for i in range(len(combined_df)):
        if math.isnan(combined_df.projected_response[i]):
            combined_df.loc[i, 'projected_response'] = combined_df.loc[i, 'response']

    # Plot
    if plot==True:
        
        sns.set(style='white', font_scale=2.2, rc={"xtick.bottom":True, "ytick.left":True})
        fig, axes = plt.subplots(figsize=(10,7))

        plt.plot(case_series_patient.day, case_series_patient.response, 'yo', linestyle='-', ms=14, mfc='white', label='SOC dosing', mew=2)
        plt.plot(combined_df[(combined_df['response']>=8) & (combined_df['response']<=10)].day, combined_df[(combined_df['response']>=8) & (combined_df['response']<=10)].response, 'yo', ms=14, label='Within the therapeutic range\nwith SOC dosing', mew=2)
        plt.plot(combined_df.day, combined_df.projected_response, 'm^', ms=14, linestyle='-', label='CURATE.AI-assisted dosing', mfc="white", mew=2)
        plt.plot(combined_df[(combined_df['projected_response']>=8) & (combined_df['projected_response']<=10)].day, combined_df[(combined_df['projected_response']>=8) & (combined_df['projected_response']<=10)].projected_response, 'm^', ms=14, label='Within the therapeutic range\nwith CURATE.AI-assisted dosing', mew=2)

        sns.despine()
        plt.xticks(np.arange(2,max(case_series_patient.day),step=4))
        plt.xlabel('Day')
        plt.ylabel('TTL (ng/ml)')
        plt.axhspan(8, 10, facecolor='grey', alpha=0.2)
        
        handles, labels = plt.gca().get_legend_handles_labels()
        order = [0,1,2,3]
        legend1 = plt.legend([handles[idx] for idx in order],[labels[idx] for idx in order],
                bbox_to_anchor=(1.04,0.54), loc='center left', frameon=False) 

        legend_elements = [Patch(facecolor='grey', edgecolor='grey',
                                label='Region within therapeutic range', alpha=.2)]
        legend2 = plt.legend(handles=legend_elements, bbox_to_anchor=(1.04,0.34), loc='upper left', frameon=False)

        axes.add_artist(legend1)
        axes.add_artist(legend2)
        
        plt.tight_layout()

        if case_series_patient_num == 5:
            plt.savefig('fig_5a.png', dpi=1000, facecolor='w', bbox_inches='tight')
            plt.savefig('fig_5a.svg', dpi=1000, facecolor='w', bbox_inches='tight')
        elif case_series_patient_num == 4:
            plt.savefig('fig_6a.png', dpi=1000, facecolor='w', bbox_inches='tight')
            plt.savefig('fig_6a.svg', dpi=1000, facecolor='w', bbox_inches='tight')


    return combined_df

def fig_5b(plot=False):
    """
    Line plot of response vs dose for patient 5's day recommendation,
    with data points as (dose, response) pairs on day 2 and 3,
    and with linear regression line. 
    """
    df = pd.read_excel(result_file, sheet_name='result')

    # Subset patient 5 and method
    df = df[(df.patient==5) & (df.method=='L_RW_wo_origin') & (df.pred_day==4)]

    df = df[['patient', 'pred_day', 'fit_dose_1', 'fit_dose_2', 'fit_response_1', 'fit_response_2', 
             'coeff_1x', 'coeff_0x', 'dose', 'response', 'prediction', 'deviation', 'abs_deviation']].reset_index(drop=True)

    # Add dose recommendation columns for tac levels of 8 to 10 ng/ml
    df['dose_recommendation_8'] = ""
    df['dose_recommendation_10'] = ""

    for i in range(len(df)):

        # Create function
        coeff = df.loc[i, 'coeff_1x':'coeff_0x'].apply(float).to_numpy()
        coeff = coeff[~np.isnan(coeff)]
        p = np.poly1d(coeff)
        x = np.linspace(0, max_dose_recommendation)
        y = p(x)

        # Check for duplicates, which will occur if coeff_1x is very close to 0, and
        # will cause RuntimeError for interp1d. Hence, set interpolated doses to the intercept,
        # also known as coeff_0x
        dupes = [x for n, x in enumerate(y) if x in y[:n]]
        if len(dupes) != 0:
            df.loc[i, 'dose_recommendation_8'] = df.loc[i, 'coeff_0x']
            df.loc[i, 'dose_recommendation_10'] = df.loc[i, 'coeff_0x']

        else:
            f = interpolate.interp1d(y, x, fill_value='extrapolate')

            df.loc[i, 'dose_recommendation_8'] = f(8)
            df.loc[i, 'dose_recommendation_10'] = f(10)

    df_original = df.copy()

    df = df[['patient', 'pred_day', 'fit_dose_1', 'fit_dose_2', 'fit_response_1', 'fit_response_2']]

    # Create dataframe for x as doses to fit regression model
    df_1 = df[['patient', 'pred_day', 'fit_dose_1', 'fit_dose_2']]
    df_1 = df_1.set_index(['patient', 'pred_day'])
    df_1 = df_1.stack().reset_index()
    df_1 = df_1.rename(columns={'level_2':'day', 0:'x'})
    df_1['day'] = df_1['day'].replace({'fit_dose_1':2, 'fit_dose_2':3})

    # Create dataframe for y as response to fit regression model
    df_2 = df[['patient', 'pred_day', 'fit_response_1', 'fit_response_2']]
    df_2 = df_2.set_index(['patient', 'pred_day'])
    df_2 = df_2.stack().reset_index()
    df_2 = df_2.rename(columns={'level_2':'day', 0:'y'})
    df_2['day'] = df_2['day'].replace({'fit_response_1':2, 'fit_response_2':3})

    combined_df = df_1.merge(df_2, how='left', on=['patient', 'pred_day', 'day'])

    if plot == True:
        # Plot
        sns.set(style='white', font_scale=2.2,
            rc={"figure.figsize":(7,7), "xtick.bottom":True, "ytick.left":True})

        # Plot regression line
        x = np.array([combined_df.x[0],combined_df.x[1]])
        y = np.array([combined_df.y[0],combined_df.y[1]])
        a, b = np.polyfit(x, y, 1)
        x_values = np.linspace(0, 3)
        plt.plot(x_values, a*x_values + b, color='y', linestyle='-', zorder=1)

        # Plot scatter points
        plt.scatter(x, y, s=200, facecolors="white", edgecolors="y", linewidths=2, zorder=2)

        # Plot therapeutic range
        plt.axhspan(8, 10, facecolor='grey', alpha=0.2)

        # Label days
        for i in range(combined_df.shape[0]):
            plt.text(x=combined_df.x[i]+0.13,y=combined_df.y[i]-0.3,s=int(combined_df.day[i]),
                    fontdict=dict(color='black',size=18),
                    bbox=dict(facecolor='none', ec='black', alpha=0.5, boxstyle='circle'))

        sns.despine()
        plt.title('Day 4 Recommendation')
        plt.xlabel('Dose (mg)')
        plt.ylabel('TTL (ng/ml)')
        plt.xticks(np.arange(0,3.5,step=0.5))
        plt.xlim(0,2.5)
        
        plt.savefig('fig_5b.png', dpi=1000, facecolor='w', bbox_inches='tight')
        plt.savefig('fig_5b.svg', dpi=1000, facecolor='w', bbox_inches='tight')

    return combined_df, df_original

# Fig 6
def fig_6a(plot=False):
    case_series_patient_num = 4
    case_series_patient_journey(case_series_patient_num=case_series_patient_num, all_data_file=all_data_file)

def fig_6b(plot=False):
    """
    Multiple plots of response vs dose for repeated dosing strategy of 
    patient 4, with each plot representing one day of prediction. 
    """

    dat_original, combined_df = fig_6_computation(result_file)

    # Subset repeated doses
    combined_df = combined_df[(combined_df.pred_day > 4) & (combined_df.pred_day <= 9)].reset_index(drop=True)

    sns.set(style='white', font_scale=2.2,
        rc={"xtick.bottom":True, "ytick.left":True})

    fig, ax = plt.subplots(1, 5, figsize=(25,5), gridspec_kw = {'wspace':0.3, 'hspace':0.5})
    # plt.subplots_adjust(wspace=-0.1)

    # Loop through number of predictions chosen
    for i in range(5):

        plt.subplot(1,5,i+1)

        # Plot regression lisne
        x = np.array([combined_df.x[i*2],combined_df.x[i*2+1]])
        y = np.array([combined_df.y[i*2],combined_df.y[i*2+1]])
        a, b = np.polyfit(x, y, 1)
        x_values = np.linspace(0, 9)
        plt.plot(x_values, a*x_values + b, linestyle='-', color='y', zorder=1)

        # Plot scatter points
        plt.scatter(x, y, s=200, facecolors="white", edgecolors="y", linewidths=2, zorder=2)

        plt.axhspan(8, 10, facecolor='grey', alpha=0.2)

        sns.despine()
        plt.ylim(0,max(combined_df.y+2))
        if i ==0:
            plt.ylabel('TTL (ng/ml)')
        plt.yticks(np.arange(0,15,step=2))
        plt.xlabel('Dose (mg)')
        plt.xticks(np.arange(0,9,step=2))
        plt.title('Day ' + str(combined_df.pred_day[i*2+1]) + '\nRecommendation')

        # Label days
        for j in range(2):
            plt.text(x=combined_df.x[i*2+j]+0.8,y=combined_df.y[i*2+j]-0.3,s=int(combined_df.day[i*2+j]), 
                fontdict=dict(color='black',size=18),
                bbox=dict(facecolor='none', ec='black', alpha=0.5, boxstyle='circle'))
            
        # Add legend for grey patch of therapeutic range
        if i == 0:
            legend_elements = [Patch(facecolor='grey', edgecolor='grey',
                                    label='Region within therapeutic range', alpha=.2)]
            plt.legend(handles=legend_elements, bbox_to_anchor=(-0.2,-.3), loc='upper left', frameon=False)       

    plt.tight_layout()
    plt.savefig('fig_6b.png',dpi=1000, bbox_inches='tight', pad_inches=0)
    plt.savefig('fig_6b.svg',dpi=1000, bbox_inches='tight', pad_inches=0)

    return combined_df

def fig_6c(plot=False):
    """Line and scatter plot of repeated dose vs day for CURATE.AI-assisted vs SOC"""
    all_data_file='all_data.xlsx'
    case_series_patient_num=4

    df = create_df_for_CURATE_assessment()
    df = df[df.patient==case_series_patient_num].reset_index(drop=True)
    df = df[(df.pred_day>=5) & (df.pred_day<=9)]

    sns.set(font_scale=2.2, rc={"figure.figsize": (7,7), "xtick.bottom":True, "ytick.left":True}, style='white')

    plt.plot(df.pred_day, df.dose, 'yo', linestyle='-', ms=14, mfc="white", label='SOC dosing', mew=2)
    plt.plot(df[(df['response']>=8) & (df['response']<=10)].pred_day, df[(df['response']>=8) & (df['response']<=10)].dose, 'yo', ms=14, label='Within the therapeutic range\nwith SOC dosing', mew=2)
    plt.plot(df.pred_day, df.dose_recommendation, 'm^', ms=14, linestyle='-', label='CURATE.AI-assisted dosing', mfc="white", mew=2)
    plt.plot(df[(df['projected_response']>=8) & (df['projected_response']<=10)].pred_day, df[(df['projected_response']>=8) & (df['projected_response']<=10)].dose_recommendation, 'm^', ms=14, label='Within the therapeutic range\nwith CURATE.AI-assisted dosing', mew=2)

    plt.legend('', frameon=False)
    sns.despine()
    plt.xlabel('Day')
    plt.ylabel('Dose (mg)')
    plt.yticks(np.arange(5,8,step=0.5))
    plt.ylim(4.5, 8)

    plt.tight_layout()
    plt.savefig('fig_6c.png',dpi=1000)
    plt.savefig('fig_6c.svg',dpi=1000)

    return df

def fig_6d(plot=True):
    """Scatter plot of dose and response for CURATE.AI-assisted and SOC dosing"""
    # dat_original, combined_df = fig_6_computation(plot=plot, dose=dose, result_file=result_file)
    # clean_dat = pd.read_excel(result_file, sheet_name='clean')

    # # Subset pred_days with repeated dose of 6mg
    # dat = dat_original[(dat_original.pred_day >= 5) & (dat_original.pred_day <= 9)].reset_index(drop=True)

    # # Add column for CURATE recommendation
    # CURATE_dosing = [7.5, 5.5, 5.5, 5, 5]
    # dat['CURATE-recommended dose'] = CURATE_dosing

    # # Add column for predicted response if CURATE dose was administered instead
    # dat['predicted_response_based_on_rec'] = dat['coeff_1x'] * dat['CURATE-recommended dose'] + dat['coeff_0x']

    # dat = dat[['pred_day', 'dose', 'response', 'CURATE-recommended dose', 'predicted_response_based_on_rec']]

    all_data_file='all_data.xlsx'
    case_series_patient_num=4

    df = create_df_for_CURATE_assessment()
    df = df[df.patient==case_series_patient_num].reset_index(drop=True)
    df = df[(df.pred_day>=5) & (df.pred_day<=9)].reset_index(drop=True)

    if plot==True:
        # Plot
        sns.set(font_scale=2.2, rc={"figure.figsize": (7,7), "xtick.bottom":True, "ytick.left":True}, style='white')
        fig, axes = plt.subplots()

        plt.plot(df['dose_recommendation'], df['projected_response'], 'm^', label='CURATE.AI-assisted dosing', ms=14, mfc="white", mew=2)
        plt.plot(df[(df['projected_response']<=therapeutic_range_upper_limit) & (df['projected_response']>=therapeutic_range_lower_limit)].dose_recommendation, \
            df[(df['projected_response']<=therapeutic_range_upper_limit) & (df['projected_response']>=therapeutic_range_lower_limit)].projected_response, \
                'm^', ms=14, mew=2)
        plt.plot(df['dose'], df['response'],'yo', label='SOC dosing', ms=14, mfc="white", mew=2)
        plt.plot(df[(df['response']<=therapeutic_range_upper_limit) & (df['response']>=therapeutic_range_lower_limit)].dose, \
            df[(df['response']<=therapeutic_range_upper_limit) & (df['response']>=therapeutic_range_lower_limit)].response, \
                'yo', ms=14, mew=2)
        sns.despine()
        plt.xlabel('Dose (mg)')
        plt.ylabel('TTL (ng/ml)')
        plt.axhspan(therapeutic_range_lower_limit, therapeutic_range_upper_limit, facecolor='grey', alpha=0.2)
        plt.xticks(np.arange(4,8.5,step=0.5))
        plt.xlim(4,8)
        plt.yticks(np.arange(8,15,step=1))

        # legend1 = plt.legend(bbox_to_anchor=(0.5,-0.3), loc='center', frameon=False)

        # legend_elements = [Patch(facecolor='grey', edgecolor='grey',
        #                         label='Region within therapeutic range', alpha=.2)]
        # legend2 = plt.legend(handles=legend_elements, bbox_to_anchor=(-0.07,-0.35), loc='upper left', frameon=False)

        # axes.add_artist(legend1)
        # axes.add_artist(legend2)        

        for i in range(df.shape[0]):
            plt.text(x=df.dose[i]+0.2,y=df.response[i],s=int(df.pred_day[i]),
                    fontdict=dict(color='black',size=18),
                    bbox=dict(facecolor='none', ec='black', alpha=0.5, boxstyle='circle'))

            plt.text(x=df.loc[i, 'dose_recommendation']+0.2,y=df.loc[i, 'projected_response'],s=int(df.pred_day[i]),
                fontdict=dict(color='black',size=18),
                bbox=dict(facecolor='none', ec='black', alpha=0.5, boxstyle='circle'))

        plt.tight_layout()
        plt.savefig('fig_6d.png',dpi=1000, bbox_inches='tight')
        plt.savefig('fig_6d.svg',dpi=1000, bbox_inches='tight')
    
    return df

def fig_6_computation(result_file):
    """
    Plot RW profiles for patient 4, with shaded region representing therapeutic range,
    colors representing prediction days, and number circles for the day from which
    the dose-response pairs were obtained from.
    """    
    dat = pd.read_excel(result_file, sheet_name='result')
    # Subset L_RW_wo_origin and patient 4
    dat = dat[(dat.method=='L_RW_wo_origin') &  (dat.patient==4)]

    dat = dat[['patient', 'method', 'pred_day', 'dose', 'response', 'coeff_1x', 'coeff_0x', 'prediction', 'deviation', 'fit_dose_1', 'fit_dose_2', 'fit_response_1', 'fit_response_2', 'day_1', 'day_2']].reset_index(drop=True)

    # Interpolate to find percentage of possible dosing events for when prediction and observed response are outside range
    for i in range(len(dat)):
        # Create function
        coeff = dat.loc[i, 'coeff_1x':'coeff_0x'].apply(float).to_numpy()
        coeff = coeff[~np.isnan(coeff)]
        p = np.poly1d(coeff)
        x = np.linspace(0, max_dose_recommendation)
        y = p(x)

        # Check for duplicates, which will occur if coeff_1x is very close to 0, and
        # will cause RuntimeError for interp1d. Hence, set interpolated doses to the intercept,
        # also known as coeff_0x
        dupes = [x for n, x in enumerate(y) if x in y[:n]]
        if len(dupes) != 0:
            dat.loc[i, 'interpolated_dose_8'] = dat.loc[i, 'coeff_0x']
            dat.loc[i, 'interpolated_dose_9'] = dat.loc[i, 'coeff_0x']
            dat.loc[i, 'interpolated_dose_10'] = dat.loc[i, 'coeff_0x']

        else:
            f = interpolate.interp1d(y, x, fill_value='extrapolate')

            dat.loc[i, 'interpolated_dose_8'] = f(8)
            dat.loc[i, 'interpolated_dose_9'] = f(9)
            dat.loc[i, 'interpolated_dose_10'] = f(10)
        
    # Create column to find points that outperform, benefit, or do not affect SOC
    dat['effect_on_SOC'] = 'none'
    dat['predict_range'] = 'therapeutic'
    dat['response_range'] = 'therapeutic'
    dat['prediction_error'] = 'acceptable'
    dat['diff_dose'] = '>0.5'
    for i in range(len(dat)):

        if (dat.prediction[i] > 10) or (dat.prediction[i] < 8):
            dat.loc[i,'predict_range'] = 'non-therapeutic'
            if (dat.response[i] > 10) or (dat.response[i] < 8):
                dat.loc[i,'response_range'] = 'non-therapeutic'
                if (round(dat.deviation[i],2) > -2) and (round(dat.deviation[i],2) < 1.5):
                    if (abs(dat.interpolated_dose_8[i] - dat.dose[i]) or abs(dat.interpolated_dose_9[i] - dat.dose[i]) or abs(dat.interpolated_dose_10[i] - dat.dose[i])) > 0.5:
                        dat.loc[i, 'effect_on_SOC'] = 'outperform'
            elif (dat.response[i] <= 10) and (dat.response[i] >= 8):
                    dat.loc[i, 'effect_on_SOC'] = 'worsen'

    dat_original = dat.copy()

    # Subset columns
    dat = dat[['pred_day', 'effect_on_SOC', 'fit_dose_1', 'fit_dose_2', 'fit_response_1', 'fit_response_2', 'day_1', 'day_2']]

    # Stack columns to fit dataframe for plotting
    df_fit_dose = dat[['pred_day', 'effect_on_SOC', 'fit_dose_1', 'fit_dose_2']]
    df_fit_dose = df_fit_dose.set_index(['pred_day', 'effect_on_SOC'])
    df_fit_dose = df_fit_dose.stack().reset_index()
    df_fit_dose.columns = ['pred_day', 'effect_on_SOC', 'fit_dose', 'x']
    df_fit_dose = df_fit_dose.reset_index()

    df_fit_response = dat[['pred_day', 'effect_on_SOC', 'fit_response_1', 'fit_response_2']]
    df_fit_response = df_fit_response.set_index(['pred_day', 'effect_on_SOC'])
    df_fit_response = df_fit_response.stack().reset_index()
    df_fit_response.columns = ['pred_day', 'effect_on_SOC', 'fit_response', 'y']
    df_fit_response = df_fit_response.reset_index()

    df_day = dat[['pred_day', 'effect_on_SOC', 'day_1', 'day_2']]
    df_day = df_day.set_index(['pred_day', 'effect_on_SOC'])
    df_day = df_day.stack().reset_index()
    df_day.columns = ['pred_day', 'effect_on_SOC', 'day_num', 'day']
    df_day = df_day.reset_index()

    combined_df = df_fit_dose.merge(df_fit_response, how='left', on=['index', 'pred_day', 'effect_on_SOC'])
    combined_df = combined_df.merge(df_day, how='left', on=['index', 'pred_day', 'effect_on_SOC'])
    
    # if plot==True:
    #     # Plot
    #     sns.set(font_scale=1.2, rc={"figure.figsize": (16,10), "xtick.bottom":True, "ytick.left":True},
    #             style='white')
    #     g = sns.lmplot(data=combined_df, x='x', y='y', hue='pred_day', ci=None, legend=False)

    #     ec = colors.to_rgba('black')
    #     ec = ec[:-1] + (0.3,)

    #     for i in range(combined_df.shape[0]):
    #         plt.text(x=combined_df.x[i]+0.3,y=combined_df.y[i]+0.3,s=int(combined_df.day[i]), 
    #         fontdict=dict(color='black',size=13),
    #         bbox=dict(facecolor='white', ec='black', alpha=0.5, boxstyle='circle'))
            
    #         plt.text(x=0+0.3,y=10.4+0.3,s=12, 
    #         fontdict=dict(color='black',size=13),
    #         bbox=dict(facecolor='white', ec=ec, boxstyle='circle'))
            
    #         plt.text(x=0+0.3,y=8.7+0.3,s=14, 
    #         fontdict=dict(color='black',size=13),
    #         bbox=dict(facecolor='white', ec=ec, boxstyle='circle'))

    #     plt.legend(bbox_to_anchor=(1.06,0.5), loc='center left', title='Day of Prediction', frameon=False)
    #     plt.xlabel('Tacrolimus dose (mg)')
    #     plt.ylabel('Tacrolimus level (ng/ml)')
    #     plt.axhspan(8, 10, facecolor='grey', alpha=0.2)

    #     # Add data point of day 12 and day 14
    #     plt.plot(0, 10.4, marker="o", markeredgecolor="black", markerfacecolor="white")
    #     plt.plot(0, 8.7, marker="o", markeredgecolor="black", markerfacecolor="white")
    #     plt.savefig('patient_4_RW_profiles_' + dose + '.png', dpi=500, facecolor='w', bbox_inches='tight')

    return dat_original, combined_df

# Effect of CURATE.AI
def effect_of_CURATE_categories():
    """
    Calculate and print out the percentage of days with projected improvement, 
    worsening, or having no effect on patient outcomes.
    """
    df = fig_7a()
    df = df.dropna().reset_index(drop=True)

    print(df.columns)

    for i in range(len(df)):
        if 'unaffected' in df['Effect of CURATE.AI-assisted dosing'][i]:
            df.loc[i, 'result'] = 'unaffected'
        elif 'improve' in df['Effect of CURATE.AI-assisted dosing'][i]:
            df.loc[i, 'result'] = 'improve'
        elif 'worsen' in df['Effect of CURATE.AI-assisted dosing'][i]:
            df.loc[i, 'result'] = 'worsen'
        else: print(f'uncertain result at index {i}')
    
    perc_of_days_improved = len(df[df['result']=='improve'])/len(df)*100
    perc_of_days_worsened = len(df[df['result']=='worsen'])/len(df)*100
    perc_of_days_unaffected = len(df[df['result']=='unaffected'])/len(df)*100

    original_stdout = sys.stdout
    with open('effect_of_CURATE_categories.txt', 'w') as f:
        sys.stdout = f
        print(f'perc_of_days_improved: {perc_of_days_improved:.2f}%, n = {len(df)}')
        print(f'perc_of_days_worsened: {perc_of_days_worsened:.2f}%, n = {len(df)}')
        print(f'perc_of_days_unaffected: {perc_of_days_unaffected:.2f}%, n = {len(df)}')
    sys.stdout = original_stdout

    return df

def effect_of_CURATE_values():
    """
    Output: 
    1) Print:
    - 1. % of days within therapeutic range
    - 2. % of participants that reach within first week
    - 3. Day where patient first achieved therapeutic range
    2) Corresponding dataframes
    """
    original_stdout = sys.stdout
    with open('effect_of_CURATE.txt', 'w') as f:
        sys.stdout = f

        df = fig_7a()

        # Drop rows where response is NaN
        df = df[df.response.notna()].reset_index(drop=True)

        # Create column of final therapeutic range result
        for i in range(len(df)):
            if 'non' in df['Effect of CURATE.AI-assisted dosing'][i]:
                df.loc[i, 'final_response_in_TR'] = False
            else:
                df.loc[i, 'final_response_in_TR'] = True

        # 1a. % of days within therapeutic range
        perc_days_within_TR = df.groupby('patient')['final_response_in_TR'].apply(lambda x: x.sum()/x.count()*100)
        perc_days_within_TR = perc_days_within_TR.reset_index(name='result')
        result_and_distribution(perc_days_within_TR.result, '1a. % of days within therapeutic range (CURATE)')


        # 1b. % of days within therapeutic range in SOC
        # Drop rows where response is NaN
        data = fig_2(plot=False)
        data = data[data.response.notna()].reset_index(drop=True)

        # Add therapeutic range column
        for i in range(len(data)):
            if (data.response[i] >= therapeutic_range_lower_limit) & (data.response[i] <= therapeutic_range_upper_limit):
                data.loc[i, 'therapeutic_range'] = True
            else:
                data.loc[i, 'therapeutic_range'] = False

        perc_therapeutic_range = data.groupby('patient')['therapeutic_range'].apply(lambda x: x.sum()/x.count()*100)
        perc_therapeutic_range = perc_therapeutic_range.to_frame().reset_index()

        # Result and distribution
        result_and_distribution(perc_therapeutic_range.therapeutic_range, '1b. % of days within therapeutic range (SOC)')


        # 1c. Comparison between 1a and 1b
        if (stats.shapiro(perc_days_within_TR.result).pvalue > 0.05) & (stats.shapiro(perc_therapeutic_range.therapeutic_range).pvalue > 0.05):
            print(f'Normal, paired, paired t-test p-value: {stats.ttest_rel(perc_days_within_TR.result, perc_therapeutic_range.therapeutic_range).pvalue:.2f}')
        else:
            print(f'Non-normal, paired, wilcoxon signed-rank test p-value: {wilcoxon(perc_days_within_TR.result, perc_therapeutic_range.therapeutic_range).pvalue:.2f}\n')

        # 2. % of participants that reach within first week
        reach_TR_in_first_week = df[df.final_response_in_TR==True].groupby('patient')['day'].first().reset_index(name='first_day')
        reach_TR_in_first_week['result'] = reach_TR_in_first_week['first_day'] <= 7
        result = reach_TR_in_first_week['result'].sum() / len(df.patient.unique()) * 100
        print(f'2. % of participants that reach within first week: {result:.2f}, {reach_TR_in_first_week["result"].sum()} out of {len(df.patient.unique())} patients\n')

        # 3a. Day where patient first achieved therapeutic range in CURATE
        result_and_distribution(reach_TR_in_first_week.first_day, '3a. Day where patient first achieved therapeutic range (CURATE)')

        # 3b. Day where patient first achieved therapeutic range in SOC
        first_TR_df = data.copy()
        first_TR_df = first_TR_df[first_TR_df['Tacrolimus trough levels (TTL)']=='Within the therapeutic range'].reset_index(drop=True)
        first_TR_df = first_TR_df.groupby('patient')['Day'].first().to_frame().reset_index()

        # Result and distribution
        result_and_distribution(first_TR_df.Day, '3b. Day where patient first achieved therapeutic range (SOC)')

        # 3c. Compare between 3a and 3b
        print(reach_TR_in_first_week.first_day)
        print(first_TR_df.Day)

        if (stats.shapiro(reach_TR_in_first_week.first_day).pvalue > 0.05) & (stats.shapiro(first_TR_df.Day).pvalue > 0.05):
            print(f'Normal, paired, paired t-test p-value: {stats.ttest_rel(reach_TR_in_first_week.first_day, first_TR_df.Day).pvalue:.2f}')
        else:
            print(f'Non-normal, paired, wilcoxon signed-rank test p-value: {wilcoxon(reach_TR_in_first_week.first_day, first_TR_df.Day).pvalue:.2f}')

    sys.stdout = original_stdout

def effect_of_CURATE_inter_indiv_differences():
    """
    1) Print the percentage of patients, out of total patients, 
    that were in therapeutic range more/less/equally frequent with CURATE.AI-assisted
    dosing. 
    2) Print the percentage of patients, out of total patients, 
    that first achieved the therapeutic range earlier/later/on the same day
    with CURATE.AI-assisted dosing.
    """
    original_stdout = sys.stdout
    with open('effect_of_CURATE_inter_indiv_differences.txt', 'w') as f:
        sys.stdout = f
        df = fig_7a().dropna()

        # Percentage of days within TTR
        SOC_TTR = df.groupby('patient')['therapeutic_range'].apply(lambda x: (x=='therapeutic').sum()/x.count()*100).reset_index(name='SOC_TTR')
        CURATE_TTR = df.groupby('patient')['Effect of CURATE.AI-assisted dosing'].apply(lambda x: (x.count()-x[x.str.contains('non-therapeutic')].count())/x.count()*100).reset_index(name='CURATE_TTR')

        combined_df_TTR = SOC_TTR.merge(CURATE_TTR, on='patient')
        for i in range(len(combined_df_TTR)):
            if combined_df_TTR.CURATE_TTR[i] > combined_df_TTR.SOC_TTR[i]:
                combined_df_TTR.loc[i, 'effect_of_CURATE_on_TTR'] = 'more frequent'
            elif combined_df_TTR.CURATE_TTR[i] < combined_df_TTR.SOC_TTR[i]:
                combined_df_TTR.loc[i, 'effect_of_CURATE_on_TTR'] = 'less frequent'
            else:
                combined_df_TTR.loc[i, 'effect_of_CURATE_on_TTR'] = 'equally frequent'

        TTR = combined_df_TTR.effect_of_CURATE_on_TTR.value_counts()/len(df.patient.unique())*100
        print(TTR)
        print(f'Out of {len(df.patient.unique())} patients')

        # First day to reach TTR
        SOC_first_day = df[df.therapeutic_range=='therapeutic'].groupby('patient')['day'].first().reset_index(name='SOC_first_day')
        CURATE_first_day = df.copy()
        CURATE_first_day = CURATE_first_day[CURATE_first_day['Effect of CURATE.AI-assisted dosing'].str.contains("non-therapeutic")==False].groupby('patient')['day'].first().reset_index(name='CURATE_first_day')

        combined_df_first_day = CURATE_first_day.merge(SOC_first_day, on='patient')
        for i in range(len(combined_df_first_day)):
            if combined_df_first_day.CURATE_first_day[i] < combined_df_first_day.SOC_first_day[i]:
                combined_df_first_day.loc[i, 'first_day_in_TTR'] = 'earlier'
            elif combined_df_first_day.CURATE_first_day[i] > combined_df_first_day.SOC_first_day[i]:
                combined_df_first_day.loc[i, 'first_day_in_TTR'] = 'later'
            else:
                combined_df_first_day.loc[i, 'first_day_in_TTR'] = 'same'

        first_day_in_TTR = combined_df_first_day.first_day_in_TTR.value_counts()/len(df.patient.unique())*100
        print(first_day_in_TTR)
        print(f'Out of {len(df.patient.unique())} patients')
    
    sys.stdout = original_stdout

# Fig 7a
def fig_7a(plot=False):
    """
    Facetgrid scatter plot of effect of CURATE on all data.

    Output:
    - Plot (saved)
    - Dataframe used to create the plot
    """

    df = create_df_for_CURATE_assessment()

    # Add column of 'Effect of CURATE.AI-assisted dosing' and 
    # add column for dose range
    df['Effect of CURATE.AI-assisted dosing'] = ""

    for i in range(len(df)):
        if (df.effect_of_CURATE[i] == 'Unaffected') & (df.therapeutic_range[i] == True):
            df.loc[i, 'Effect of CURATE.AI-assisted dosing'] = 'Unaffected, remain as therapeutic range'
        elif (df.effect_of_CURATE[i] == 'Unaffected') & (df.therapeutic_range[i] == False):
            df.loc[i, 'Effect of CURATE.AI-assisted dosing'] = 'Unaffected, remain as non-therapeutic range'
        elif df.effect_of_CURATE[i] == 'Improve':
            df.loc[i, 'Effect of CURATE.AI-assisted dosing'] = 'Improve to therapeutic range'
        else:
            df.loc[i, 'Effect of CURATE.AI-assisted dosing'] = 'Worsen to non-therapeutic range'

    # Rename column in df from 'pred_day' to 'day' before merging
    df = df.rename(columns={'pred_day':'day'})

    # Subset columns
    df = df[['patient', 'day', 'Effect of CURATE.AI-assisted dosing']]

    # Import all data
    all_data = pd.read_excel(all_data_file)

    # Subset patient, day, response, dose 
    all_data = all_data[['patient', 'day', 'response', 'dose']]

    # Add therapeutic range column
    all_data['therapeutic_range'] = ""
    for i in range(len(all_data)):
        if (all_data.response[i] >= therapeutic_range_lower_limit) & (all_data.response[i] <= therapeutic_range_upper_limit):
            all_data.loc[i, 'therapeutic_range'] = 'therapeutic'
        else:
            all_data.loc[i, 'therapeutic_range'] = 'non_therapeutic'

    # Merge on all columns in all_data
    combined_dat = all_data.merge(df, how='left', on=['patient', 'day'])

    # Add dose range column and fill in 'Effect of CURATE'
    combined_dat['Dose range'] = ""

    for i in range(len(combined_dat)):
        if np.isnan(combined_dat.dose[i]):
            combined_dat.loc[i, 'Dose range'] = 'Unavailable'
        elif combined_dat.dose[i] < low_dose_upper_limit:
            combined_dat.loc[i, 'Dose range'] = 'Low'
        elif combined_dat.dose[i] < medium_dose_upper_limit:
            combined_dat.loc[i, 'Dose range'] = 'Medium'
        else: 
            combined_dat.loc[i, 'Dose range'] = 'High'

        if str(combined_dat['Effect of CURATE.AI-assisted dosing'][i])=='nan':
            if combined_dat.therapeutic_range[i] == 'therapeutic':
                combined_dat.loc[i, 'Effect of CURATE.AI-assisted dosing'] = 'Unaffected, remain as therapeutic range'
            else:
                combined_dat.loc[i, 'Effect of CURATE.AI-assisted dosing'] = 'Unaffected, remain as non-therapeutic range' 

    # Rename effect of CURATE
    plot_df = combined_dat.copy()
    plot_df['Effect of CURATE.AI-assisted dosing'] = plot_df['Effect of CURATE.AI-assisted dosing'].map({\
        'Unaffected, remain as therapeutic range':'TTL is unaffected (remains within the therapeutic range)',
        'Unaffected, remain as non-therapeutic range':'TTL is unaffected (remains outside of the therapeutic range)',
        'Improve to therapeutic range':'TTL improves (moves from outside to within the therapeutic range)',
        'Worsen to non-therapeutic range':'TTL worsens (moves from within to outside the therapeutic range)'})
    
    if plot == True:
        # Plot
        sns.set(font_scale=1.5, rc={"figure.figsize": (16,10), "xtick.bottom":True, "ytick.left":True}, style='white')
        hue_order = ['TTL is unaffected (remains within the therapeutic range)', 'TTL is unaffected (remains outside of the therapeutic range)',
                    'TTL improves (moves from outside to within the therapeutic range)', 'TTL worsens (moves from within to outside the therapeutic range)']
        palette = [sns.color_palette()[1], sns.color_palette()[0], sns.color_palette()[2],\
                sns.color_palette()[3]]
        style_order = ['Low', 'Medium', 'High', 'Unavailable']

        # Scatter point
        g = sns.relplot(data=plot_df, x='day', y='response', hue='Effect of CURATE.AI-assisted dosing',\
                        hue_order=hue_order, col='patient', palette=palette,\
                        col_wrap=4, height=3, aspect=1, s=100, style_order=style_order, zorder=2, edgecolor=None)

        # Move legend below plot
        sns.move_legend(g, 'upper left', bbox_to_anchor=(0,0), \
            title='Effect of CURATE.AI-assisted dosing on tacrolimus trough levels (TTL)', ncol=1)

        # Titles and labels
        g.set_titles('Patient {col_name}')
        g.set(yticks=np.arange(0,math.ceil(max(plot_df['response'])),4),
            xticks=np.arange(0,max(plot_df.day),step=5))
        g.set_ylabels('TTL (ng/ml)')
        g.set_xlabels('Day')

        # Add gray region for therapeutic range
        for ax in g.axes:
            ax.axhspan(therapeutic_range_lower_limit, therapeutic_range_upper_limit, facecolor='grey', alpha=0.2, zorder=1)

        legend1 = plt.legend()
        legend_elements = [Patch(facecolor='grey', edgecolor='grey',
                            label='Region within therapeutic range', alpha=.2)]
        legend2 = plt.legend(handles=legend_elements, bbox_to_anchor=(-0.5,-0.42), loc='upper left', frameon=False)

        # plt.show()
        # plt.tight_layout()
        plt.savefig('fig_7a.png', dpi=1000, facecolor='w', bbox_inches='tight')
        plt.savefig('fig_7a.svg', dpi=1000, facecolor='w', bbox_inches='tight')

    return plot_df

# Fig 7b
def fig_7b(plot=False):
    """
    Boxplot for day when TR is first achieved, for
    both SOC and CURATE
    """
    plot_df = fig_7b_computation()

    if plot == True:
        
        sns.set(font_scale=1.2, rc={"figure.figsize": (5,5), "xtick.bottom":True, "ytick.left":True}, style='white')
        fig, ax = plt.subplots()

        # Boxplot
        g = sns.boxplot(x='Dosing', y='First day in therapeutic range', data=plot_df, width=0.5, zorder=1)
        g.set_xlabel(None)
        g.set_xticklabels(['SOC dosing\n(N = 15)', 'CURATE.AI-assisted\ndosing\n(N = 15)'])
        
        # Scatter points
        SOC_df = plot_df[plot_df.Dosing=='SOC'].reset_index(drop=True)['First day in therapeutic range']
        CURATE_df = plot_df[plot_df.Dosing=='CURATE'].reset_index(drop=True)['First day in therapeutic range']
        
        plt.scatter(np.zeros(len(SOC_df)), SOC_df, c='k', zorder=2)
        plt.scatter(np.ones(len(CURATE_df)), CURATE_df, c='k', zorder=3)
        for i in range(len(SOC_df)):
            plt.plot([0,1], [SOC_df[i], CURATE_df[i]], c='k', alpha=.5)

        # Aesthetics
        sns.despine()

        # Bracket and star
        x1, x2 = 0, 1
        y, h = SOC_df.max() + 3.5, 1
        plt.plot([0, 0, 1, 1], [y, y+h, y+h, y], lw=1.5, c='k')
        plt.text((x1+x2)*.5, y+h, "*", ha='center', va='bottom', color='k')    

        # Box labels
        # rects = ax.patches
        medians = [median(SOC_df), median(CURATE_df)]
        lower_quartile = [SOC_df.quantile(0.25), CURATE_df.quantile(0.25)]
        upper_quartile = [SOC_df.quantile(0.75), CURATE_df.quantile(0.75)]
        labels = [f"{i:.2f}\n({j:.2f} - {k:.2f})" for i,j,k in zip(medians, lower_quartile, upper_quartile)]
        
        plt.text(0, SOC_df.max()+0.9, labels[0], ha='center', va='bottom', 
                color='k', fontsize=13)
        plt.text(1, CURATE_df.max()+1.5, labels[1], ha='center', va='bottom', 
                color='k', fontsize=13)

        # Save
        plt.tight_layout()
        plt.savefig('fig_7b.png', dpi=1000)
        plt.savefig('fig_7b.svg', dpi=1000)

def fig_7b_computation():
        # SOC
    SOC = fig_2(plot=False)
    SOC = SOC[SOC.response.notna()].reset_index(drop=True)

    # Add therapeutic range column
    for i in range(len(SOC)):
        if (SOC.response[i] >= therapeutic_range_lower_limit) & (SOC.response[i] <= therapeutic_range_upper_limit):
            SOC.loc[i, 'therapeutic_range'] = True
        else:
            SOC.loc[i, 'therapeutic_range'] = False

    SOC = SOC[SOC['Tacrolimus trough levels (TTL)']=='Within the therapeutic range'].reset_index(drop=True)
    SOC = SOC.groupby('patient')['Day'].first().reset_index(name='SOC')

    # CURATE
    CURATE = fig_7a()

    # Drop rows where response is NaN
    CURATE = CURATE[CURATE.response.notna()].reset_index(drop=True)

    # Create column of final therapeutic range result
    for i in range(len(CURATE)):
        if ('remains outside' in CURATE['Effect of CURATE.AI-assisted dosing'][i]) | ('moves from within to outside' in CURATE['Effect of CURATE.AI-assisted dosing'][i]):
            CURATE.loc[i, 'final_response_in_TR'] = False
        else:
            CURATE.loc[i, 'final_response_in_TR'] = True

    CURATE = CURATE[CURATE.final_response_in_TR==True].groupby('patient')['day'].first().reset_index(name='CURATE')

    # Merge SOC and CURATE into one dataframe
    combined_df = SOC.merge(CURATE, how='left', on='patient')

    # Compare medians
    print(combined_df)

    if (stats.shapiro(combined_df.SOC).pvalue < 0.05) or (stats.shapiro(combined_df.CURATE).pvalue < 0.05):
        print(f'Non-normal distribution, Wilcoxon p value: {wilcoxon(combined_df.SOC, combined_df.CURATE).pvalue:.2f}')
    else:
        print(f'Normal distribution, Paired t test p value: {stats.ttest_rel(combined_df.SOC, combined_df.CURATE).pvalue:.2f}')

    # Rearrange dataframe for seaborn boxplot

    plot_df = combined_df.set_index('patient')
    plot_df = plot_df.stack().reset_index()

    plot_df = plot_df.rename(columns={'level_1':'Dosing', 0:'First day in therapeutic range'})
    plot_df['First day in therapeutic range'] = plot_df['First day in therapeutic range'].astype(float)

    return plot_df

# Fig 7c
def fig_7c(plot=False):
    """
    Barplot of % of patients in TR within first week, of
    SOC and CURATE.
    """
    # SOC
    data = fig_2(plot=False)

    # Drop rows where response is NaN
    data = data[data.response.notna()].reset_index(drop=True)

    # Add therapeutic range column
    for i in range(len(data)):
        if (data.response[i] >= therapeutic_range_lower_limit) & (data.response[i] <= therapeutic_range_upper_limit):
            data.loc[i, 'therapeutic_range'] = True
        else:
            data.loc[i, 'therapeutic_range'] = False

    first_week_df = data.copy()
    first_week_df = first_week_df[first_week_df['Tacrolimus trough levels (TTL)']=='Within the therapeutic range'].reset_index(drop=True)
    first_week_df = (first_week_df.groupby('patient')['Day'].first() <= 7).to_frame().reset_index()
    result = first_week_df.Day.sum()/first_week_df.Day.count()*100

    SOC = result

    # CURATE
    df = fig_7a()
    # Drop rows where response is NaN
    df = df[df.response.notna()].reset_index(drop=True)

    # Create column of final therapeutic range result
    for i in range(len(df)):
        if ('remains outside' in df['Effect of CURATE.AI-assisted dosing'][i]) | ('moves from within to outside' in df['Effect of CURATE.AI-assisted dosing'][i]):
            df.loc[i, 'final_response_in_TR'] = False
        else:
            df.loc[i, 'final_response_in_TR'] = True

    # % of participants that reach within first week
    reach_TR_in_first_week = df[df.final_response_in_TR==True].groupby('patient')['day'].first().reset_index(name='first_day')
    reach_TR_in_first_week['result'] = reach_TR_in_first_week['first_day'] <= 7

    result = reach_TR_in_first_week['result'].sum() / len(reach_TR_in_first_week) * 100

    CURATE = result

    # Restructure dataframe for plotting with seaborn
    plot_df = pd.DataFrame({'SOC':[SOC], 'CURATE':[CURATE]}).stack()
    plot_df = plot_df.to_frame().reset_index()
    plot_df = plot_df.rename(columns={'level_1':'Dosing', 0:'perc_reach_TR_in_first_week'})

    # Plot
    if plot == True:
        sns.set(font_scale=1.2, rc={"figure.figsize": (5,5), "xtick.bottom":True, "ytick.left":True}, style='white')
        fig, ax = plt.subplots()
        p = ax.bar(plot_df.Dosing, plot_df.perc_reach_TR_in_first_week, width=.5, 
        color=[sns.color_palette("Paired",10)[8],sns.color_palette("Paired",10)[9]], edgecolor='k')
        
        # Aesthetics
        sns.despine()
        ax.set_xticklabels(['SOC dosing\n(N = 15)', 'CURATE.AI-assisted\ndosing\n(N = 15)'])
        plt.ylabel('Patients who achieve therapeutic\nrange in first week (%)')
        
        # Bar labels
        ax.bar_label(p, fmt='%.2f', fontsize=13)
        # plt.show()
        plt.savefig('fig_7c.png', dpi=1000, facecolor='w', bbox_inches='tight')
        plt.savefig('fig_7c.svg', dpi=1000, facecolor='w', bbox_inches='tight')

    return plot_df

# Fig 7d
def SOC_CURATE_perc_in_TR():
    """
    Boxplot of % of days in TR, for SOC and CURATE.
    Print out kruskal wallis test for difference in medians.
    """

    # SOC
    perc_days_TR_SOC = fig_2(plot=False)

    # Drop rows where response is NaN
    perc_days_TR_SOC = perc_days_TR_SOC[perc_days_TR_SOC.response.notna()].reset_index(drop=True)

    # Add therapeutic range column
    for i in range(len(perc_days_TR_SOC)):
        if (perc_days_TR_SOC.response[i] >= therapeutic_range_lower_limit) & (perc_days_TR_SOC.response[i] <= therapeutic_range_upper_limit):
            perc_days_TR_SOC.loc[i, 'therapeutic_range'] = True
        else:
            perc_days_TR_SOC.loc[i, 'therapeutic_range'] = False

    perc_days_TR_SOC = perc_days_TR_SOC.groupby('patient')['therapeutic_range'].apply(lambda x: x.sum()/x.count()*100)
    perc_days_TR_SOC = perc_days_TR_SOC.reset_index(name='SOC')

    # CURATE
    perc_days_TR_CURATE = fig_7a()

    # Drop rows where response is NaN
    perc_days_TR_CURATE = perc_days_TR_CURATE[perc_days_TR_CURATE.response.notna()].reset_index(drop=True)

    # Create column of final therapeutic range result
    for i in range(len(perc_days_TR_CURATE)):

        if ('remains outside' in perc_days_TR_CURATE['Effect of CURATE.AI-assisted dosing'][i]) | ('moves from within to outside' in perc_days_TR_CURATE['Effect of CURATE.AI-assisted dosing'][i]):
        # if 'non' in perc_days_TR_CURATE['Effect of CURATE.AI-assisted dosing'][i]:
            perc_days_TR_CURATE.loc[i, 'final_response_in_TR'] = False
        else:
            perc_days_TR_CURATE.loc[i, 'final_response_in_TR'] = True

    perc_days_TR_CURATE = perc_days_TR_CURATE.groupby('patient')['final_response_in_TR'].apply(lambda x: x.sum()/x.count()*100)
    perc_days_TR_CURATE = perc_days_TR_CURATE.reset_index(name='CURATE')

    perc_days_TR = perc_days_TR_SOC.merge(perc_days_TR_CURATE, how='left', on='patient')

    # Compare medians
    print('Comparison of medians between SOC and CURATE\n')
    if (stats.shapiro(perc_days_TR.SOC).pvalue < 0.05) or (stats.shapiro(perc_days_TR.CURATE).pvalue < 0.05):
        print(f'Non-normal distribution, Wilcoxon p value: {wilcoxon(perc_days_TR.SOC, perc_days_TR.CURATE).pvalue:.2f}')
    else:
        print(f'Normal distribution, Paired t test p value: {stats.ttest_rel(perc_days_TR.SOC, perc_days_TR.CURATE).pvalue:.2f}')

    # Rearrange dataframe for seaborn boxplot
    perc_days_TR_plot = perc_days_TR.set_index('patient')
    perc_days_TR_plot = perc_days_TR_plot.stack().reset_index()
    perc_days_TR_plot = perc_days_TR_plot.rename(columns={'level_1':'Dosing', 0:'Days in therapeutic range (%)'})
    
    return perc_days_TR

def fig_7d():
    df = SOC_CURATE_perc_in_TR()

    fig, ax = plt.subplots()
    # plt.figure(figsize=(5,5))
    sns.set(font_scale=1.2, rc={"xtick.bottom":True, "ytick.left":True}, style='white')

    # Bar plot
    p = ax.bar(['SOC dosing\n(N = 16)', 'CURATE.AI-assisted\ndosing\n(N = 16)'], [mean(df.SOC), mean(df.CURATE)], yerr=[stdev(df.SOC), stdev(df.CURATE)],
        edgecolor='black', capsize=10, color=[sns.color_palette("Paired",8)[0],sns.color_palette("Paired",8)[1]], zorder=1, width=.4)

    # Scatter points
    plt.scatter(np.zeros(len(df.SOC)), df.SOC, c='k', zorder=2)
    plt.scatter(np.ones(len(df.CURATE)), df.CURATE, c='k', zorder=3)
    for i in range(len(df.CURATE)):
        plt.plot([0,1], [df.SOC[i], df.CURATE[i]], c='k', alpha=.5)

    # Aesthetics
    plt.ylabel('Days in therapeutic range (%)')
    sns.despine()
    plt.ylim(0, 65)

    # Bar labels
    rects = ax.patches
    averages = [mean(df.SOC), mean(df.CURATE)]
    SD = [stdev(df.SOC), stdev(df.CURATE)]
    labels = [f"{i:.2f} ± {j:.2f}" for i,j in zip(averages, SD)]
    for rect, label in zip(rects, labels):
        height = rect.get_height()
        ax.text(
            rect.get_x() + rect.get_width() / 2, height + 35, label, 
            ha="center", va="bottom", fontsize=13
        )

    plt.tight_layout()
    plt.savefig('fig_7d.png', dpi=1000, facecolor='w', bbox_inches='tight')
    plt.savefig('fig_7d.svg', dpi=1000, facecolor='w', bbox_inches='tight')

    return df

# Assessment of CURATE.AI
def create_df_for_CURATE_assessment():
    file_name = 'dose_recommendations.xlsx'

    dat = pd.read_excel(file_name)

    CURATE_assessment = dat[['patient', 'pred_day', 'prediction', 'response', 'deviation', 'dose', 'dose_recommendation', 'predicted_response_after_recommended_dose']]

    # Add columns for assessment
    CURATE_assessment['reliable'] = ""
    CURATE_assessment['accurate'] = ""
    CURATE_assessment['diff_dose'] = ""
    CURATE_assessment['therapeutic_range'] = ""
    CURATE_assessment['actionable'] = ""
    CURATE_assessment['effect_of_CURATE'] = ""

    for i in range(len(CURATE_assessment)):

        # Reliable
        if (CURATE_assessment.prediction[i] >= therapeutic_range_lower_limit) & (CURATE_assessment.prediction[i] <= therapeutic_range_upper_limit):
            prediction_range = 'therapeutic'
        else:
            prediction_range = 'non-therapeutic'

        if (CURATE_assessment.response[i] >= therapeutic_range_lower_limit) & (CURATE_assessment.response[i] <= therapeutic_range_upper_limit):
            response_range = 'therapeutic'
        else:
            response_range = 'non-therapeutic'

        reliable = (prediction_range == response_range)
        CURATE_assessment.loc[i, 'reliable'] = reliable

        # Accurate
        accurate = (dat.deviation[i] >= overprediction_limit) & (dat.deviation[i] <= underprediction_limit)
        CURATE_assessment.loc[i, 'accurate'] = accurate

        # Different dose
        diff_dose = (dat.dose[i] != dat.dose_recommendation[i])
        CURATE_assessment.loc[i, 'diff_dose'] = diff_dose

        # Therapeutic range
        therapeutic_range = (CURATE_assessment.response[i] >= therapeutic_range_lower_limit) & (CURATE_assessment.response[i] <= therapeutic_range_upper_limit)
        CURATE_assessment.loc[i, 'therapeutic_range'] = therapeutic_range

        # Actionable
        actionable = ((dat.dose_recommendation[i]) <= max_dose_recommendation) & ((dat.dose_recommendation[i]) >= min_dose_recommendation)
        CURATE_assessment.loc[i, 'actionable'] = actionable

        # Effect of CURATE
        if (reliable == True) & (accurate == True) & (diff_dose == True) & (therapeutic_range == False) & (actionable == True):
            CURATE_assessment.loc[i, 'effect_of_CURATE'] = 'Improve'
        elif (reliable == True) & (accurate == False) & (diff_dose == True) & (therapeutic_range == True) & (actionable == True):
            CURATE_assessment.loc[i, 'effect_of_CURATE'] = 'Worsen'
        elif (reliable == False) & (accurate == True) & (diff_dose == True) & (therapeutic_range == True) & (actionable == True):
            CURATE_assessment.loc[i, 'effect_of_CURATE'] = 'Worsen'
        elif (reliable == False) & (accurate == False) & (diff_dose == True) & (therapeutic_range == True) & (actionable == True):
            CURATE_assessment.loc[i, 'effect_of_CURATE'] = 'Worsen'
        else:
            CURATE_assessment.loc[i, 'effect_of_CURATE'] = 'Unaffected'

        # Compute the projected response based on Supplementary Fig S1
        for i in range(len(CURATE_assessment)):
            if CURATE_assessment.effect_of_CURATE[i] == 'Unaffected':
                if reliable == True:
                    if actionable == True:
                        CURATE_assessment.loc[i, 'projected_response'] = CURATE_assessment.loc[i, 'response']
                    else:
                        CURATE_assessment.loc[i, 'projected_response'] = CURATE_assessment.loc[i, 'predicted_response_after_recommended_dose']
                else:
                    if actionable == False:
                        CURATE_assessment.loc[i, 'projected_response'] = CURATE_assessment.loc[i, 'response']
                    else:
                        CURATE_assessment.loc[i, 'projected_response'] = 7


                CURATE_assessment.loc[i, 'projected_response'] = CURATE_assessment.loc[i, 'response']
            
            elif CURATE_assessment.effect_of_CURATE[i] == 'Improve':
                CURATE_assessment.loc[i, 'projected_response'] = CURATE_assessment.loc[i, 'predicted_response_after_recommended_dose']
            # Worsen
            else:
                # Worsen
                CURATE_assessment.loc[i, 'projected_response'] = 7

    return CURATE_assessment

# Create lists
def find_list_of_patients():
    # Declare list
    list_of_patients = []

    # Create list of patients
    wb = load_workbook(raw_data_file, read_only=True)
    list_of_patients = wb.sheetnames

    return list_of_patients

# Import data
def import_raw_data_including_non_ideal():
    df = pd.read_excel('all_data_including_non_ideal.xlsx', sheet_name='data')

    return df

# Statistical test
def result_and_distribution(df, metric_string):
    """
    Determine which normality test to do based on n.
    Conduct normality test and print results.
    Based on p-value of normality test, choose mean/median to print.
    """
    p = shapiro_test_result(df, metric_string)
        
    if p < 0.05:
        median_IQR_range(df)
    else:
        mean_and_SD(df)
        
def shapiro_test_result(df, metric_string):
    shapiro_result = stats.shapiro(df).pvalue
    
    if shapiro_result < 0.05:
        result_string = 'reject normality'
    else:
        result_string = 'assume normality'

    print(f'{metric_string}:\nShapiro test p-value = {shapiro_result:.2f}, {result_string}')
    
    return shapiro_result
    
def mean_and_SD(df):
    # Mean and SD
    mean = df.describe()['mean']
    std = df.describe()['std']
    count = df.describe()['count']

    print(f'mean {mean:.2f} | std {std:.2f} | count {count}\n')
    
def median_IQR_range(df):
    median = df.describe()['50%']
    lower_quartile = df.describe()['25%']
    upper_quartile = df.describe()['75%']
    minimum = df.describe()['min']
    maximum = df.describe()['max']
    count = df.describe()['count']

    print(f'median {median:.2f} | IQR {lower_quartile:.2f} - {upper_quartile:.2f} | count {count} | range {minimum:.2f} - {maximum:.2f}\n')

if __name__ == '__main__':
    # Get arguments
    import argparse
    parser = argparse.ArgumentParser(description='Plot CURATE.AI results')
    parser.add_argument("-f", "--figure", type=str, default=False)
    parser.add_argument("-a", "--analysis", type=str, default=False)
    args = parser.parse_args()
    
    # Figures
    if args.figure=='fig_2':
        fig_2(plot=True)
    elif args.figure=='fig_5':
        fig_5a(plot=True)
        fig_5b(plot=True)
    elif args.figure=='fig_6':
        fig_6a(plot=True)
        fig_6b(plot=True)
        fig_6c(plot=True)
        fig_6d(plot=True)
    elif args.figure=='fig_7':
        fig_7a(plot=True)
        fig_7b(plot=True)
        fig_7c(plot=True)
        fig_7d()
        
    else:
        print('no valid figure was specified')

    # Analysis
    if args.analysis=='patient_population':
        percentage_of_pts_that_reached_TR_per_dose_range()
        patient_population_values()
    elif args.analysis=='technical_perf_metrics':
        technical_performance_metrics()
    elif args.analysis=='clinically_relevant_perf_metrics':
        clinically_relevant_performance_metrics()
    elif args.analysis=='effect_of_CURATE':
        effect_of_CURATE_categories()
        effect_of_CURATE_values()
        effect_of_CURATE_inter_indiv_differences()
    elif args.analysis=='fig_4_values':
        values_in_clinically_relevant_flow_chart()
    else:
        print('no valid analysis was specified')